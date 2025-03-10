"""
Final HTML Export Module for Blockchain Analysis Report
Optimized for size and consistent column alignment with improved formatting.
"""
import os
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def get_cell_value(cell: Cell, is_block_number: bool = False) -> str:
    """Extract and format cell value, handling different types."""
    if cell.value is None:
        return ""
    
    # Handle numeric values with specific formatting
    if is_block_number and isinstance(cell.value, (int, float)):
        # Plain integer format for block numbers
        return f"{int(cell.value)}"
    elif cell.number_format.endswith('%') and isinstance(cell.value, (int, float)):
        return f"{cell.value:.1%}"
    elif isinstance(cell.value, (int, float)):
        return f"{cell.value:,.4f}" if cell.value % 1 != 0 else f"{cell.value:,.0f}"
    
    return str(cell.value)


def get_cell_hyperlink(cell: Cell) -> Optional[str]:
    """Extract hyperlink from cell if present."""
    return cell.hyperlink.target if cell.hyperlink else None


def get_cell_style(cell: Cell) -> Dict:
    """Extract cell styling information."""
    style = {}
    
    # Get text color
    if cell.font and cell.font.color:
        if cell.font.color.rgb:
            rgb = cell.font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:  # ARGB format
                style['color'] = f"#{rgb[2:]}"  # Remove alpha channel
    
    # Get bold/italic status
    if cell.font:
        if cell.font.bold:
            style['font-weight'] = 'bold'
        if cell.font.italic:
            style['font-style'] = 'italic'
        if cell.font.underline:
            style['text-decoration'] = 'underline'
    
    # Get background color
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        rgb = cell.fill.start_color.rgb
        if isinstance(rgb, str) and len(rgb) == 8:  # ARGB format
            if rgb != 'FF000000':  # Not default black
                style['background-color'] = f"#{rgb[2:]}"  # Remove alpha channel
    
    # Get alignment
    if cell.alignment:
        if cell.alignment.horizontal:
            style['text-align'] = cell.alignment.horizontal
        if cell.alignment.vertical:
            style['vertical-align'] = cell.alignment.vertical
    
    # Get borders
    if cell.border:
        for side in ['top', 'right', 'bottom', 'left']:
            border = getattr(cell.border, side)
            if border and border.style and border.style != 'none':
                style[f'border-{side}'] = '1px solid #D3D3D3'
    
    return style


def extract_worksheet_data(ws: Worksheet, keep_columns: List[str] = None, 
                          max_rows_per_sheet: int = None) -> Tuple[Dict, List[Dict]]:
    """
    Extract data from Excel worksheet into structured format for HTML export.
    
    Args:
        ws: OpenPyXL worksheet object
        keep_columns: List of column headers to include (if None, include all)
        max_rows_per_sheet: Maximum number of rows to extract per sheet (for size control)
    
    Returns:
        Tuple containing (summary_data, main_table_data)
    """
    # Skip Dashboard and Leaderboard tabs
    if ws.title in ['Dashboard', 'Leaderboard']:
        return None, None
    
    # First, extract the summary data (C2:H7)
    summary_data = {
        'address': ws['C2'].value,
        'summary_rows': []
    }
    
    # Extract summary rows (4-7)
    for row in range(4, 8):
        # Special handling for Start Block and End Block - remove commas
        is_block_cell = ws[f'C{row}'].value in ['Start Block:', 'End Block:']
        
        summary_data['summary_rows'].append({
            'label': get_cell_value(ws[f'C{row}']),
            'value': get_cell_value(ws[f'D{row}'], is_block_cell),
            'label_style': get_cell_style(ws[f'C{row}']),
            'value_style': get_cell_style(ws[f'D{row}']),
            'extra_label': get_cell_value(ws[f'G{row}']),
            'extra_value': get_cell_value(ws[f'H{row}']),
            'extra_label_style': get_cell_style(ws[f'G{row}']),
            'extra_value_style': get_cell_style(ws[f'H{row}'])
        })
    
    # Now extract the header row (row 9)
    headers = []
    header_indices = {}  # Map header names to column indices
    
    # First pass: identify all headers and their column indices
    for col_idx in range(2, ws.max_column + 1):  # Start from column B (index 2)
        cell = ws.cell(row=9, column=col_idx)
        header_text = get_cell_value(cell)
        if header_text:
            # Skip dateTime column
            if header_text == 'dateTime':
                continue
                
            # Rename 'Balance (ETH + WETH)' to 'Balance'
            if header_text == 'Balance (ETH + WETH)':
                header_text = 'Balance'
            header_indices[header_text] = col_idx
    
    # Second pass: keep only the specified columns in the right order
    if keep_columns:
        for header_name in keep_columns:
            if header_name == 'dateTime':  # Skip dateTime
                continue
                
            # Handle the renamed column
            lookup_name = 'Balance (ETH + WETH)' if header_name == 'Balance' else header_name
            if lookup_name in header_indices:
                col_idx = header_indices[lookup_name]
                cell = ws.cell(row=9, column=col_idx)
                headers.append({
                    'text': header_name,  # Use the new name
                    'style': get_cell_style(cell),
                    'column': col_idx
                })
    else:
        # If no columns specified, keep all in original order (except dateTime)
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=9, column=col_idx)
            header_text = get_cell_value(cell)
            if header_text and header_text != 'dateTime':
                # Rename 'Balance (ETH + WETH)' to 'Balance'
                if header_text == 'Balance (ETH + WETH)':
                    header_text = 'Balance'
                headers.append({
                    'text': header_text,
                    'style': get_cell_style(cell),
                    'column': col_idx
                })
    
    # Extract data rows
    data_rows = []
    
    # Determine the range of rows to extract
    max_row = ws.max_row
    if max_rows_per_sheet and max_rows_per_sheet < (max_row - 9):
        max_row = 9 + max_rows_per_sheet
    
    for row_idx in range(10, max_row + 1):
        row_data = []
        row_profit = None
        
        # First find the profit for sorting later
        for header in headers:
            if header['text'] == 'Profit':
                profit_cell = ws.cell(row=row_idx, column=header['column'])
                if profit_cell.value and isinstance(profit_cell.value, (int, float)):
                    row_profit = float(profit_cell.value)
                break
        
        # Then extract all cell data
        for header in headers:
            col_idx = header['column']
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Special handling for block numbers
            is_block_number = header['text'] == 'block #'
            
            cell_data = {
                'text': get_cell_value(cell, is_block_number),
                'style': get_cell_style(cell),
                'hyperlink': get_cell_hyperlink(cell)
            }
            
            # Special handling for specific columns
            header_name = header['text']
            
            # For links like tx Hash, shorten display text but keep full link
            if header_name in ['tx Hash']:
                if cell_data['hyperlink'] and cell_data['text']:
                    if len(cell_data['text']) > 12:
                        # Save full hash but display shortened
                        cell_data['full_text'] = cell_data['text']
                        cell_data['text'] = cell_data['text'][:10] + '...'
            
            # Format Profit values with appropriate styling
            if header_name == 'Profit' and cell_data['text'] and cell_data['text'] != 'None':
                try:
                    value = float(cell_data['text'].replace(',', ''))
                    if value > 0:
                        cell_data['class'] = 'positive'
                    elif value < 0:
                        cell_data['class'] = 'negative'
                except ValueError:
                    pass
            
            # Add monospace class for code-like fields
            if header_name in ['block #', 'tx Hash', 'tx Index']:
                cell_data['class'] = cell_data.get('class', '') + ' monospace'
            
            row_data.append(cell_data)
        
        # Add profit for sorting
        data_rows.append({'profit': row_profit, 'data': row_data})
    
    # Sort rows by profit (descending)
    data_rows.sort(key=lambda x: x['profit'] if x['profit'] is not None else float('-inf'), reverse=True)
    
    # Extract only the row data after sorting
    sorted_rows = [row['data'] for row in data_rows]
    
    # Combine into main table data
    main_table_data = {
        'headers': headers,
        'rows': sorted_rows
    }
    
    return summary_data, main_table_data


def generate_html_report(excel_file: str, output_file: str = None, 
                         columns_to_keep: List[str] = None,
                         max_rows_per_sheet: int = 100) -> str:
    """
    Generate a single HTML file from an Excel blockchain analysis report.
    
    Args:
        excel_file: Path to the Excel file
        output_file: Path to save the HTML output (if None, return as string)
        columns_to_keep: List of column headers to include
        max_rows_per_sheet: Maximum number of rows to include per sheet (for size control)
    
    Returns:
        HTML content as string if output_file is None, otherwise None
    """
    if columns_to_keep is None:
        columns_to_keep = [
            'block #', 'tx Hash', 'tx Index', 'miner',
            'Balance', 'Total Tip', 'Extractable Value', 'Profit',
            'Bribe %', 'Profit Rank', 'Resim', 'Resim @ txIx 0', 'Resim @ txIx -1'
        ]
    
    wb = load_workbook(excel_file, data_only=True)  # data_only=True to get values instead of formulas
    
    # Extract data from each worksheet
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Dashboard', 'Leaderboard']:
            continue  # Skip these tabs
            
        ws = wb[sheet_name]
        summary_data, main_table_data = extract_worksheet_data(ws, columns_to_keep, max_rows_per_sheet)
        
        if summary_data and main_table_data:  # Only add if data was successfully extracted
            sheets_data[sheet_name] = {
                'summary': summary_data,
                'main_table': main_table_data
            }
    
    # Get timestamp for the report
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Generate HTML
    html_content = generate_html_template(sheets_data, timestamp)
    
    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        return None
    else:
        return html_content


def generate_html_template(sheets_data: Dict, timestamp: str) -> str:
    """Generate the complete HTML template with minimal external dependencies."""
    
    # Create a minimal HTML file that loads libraries from CDN
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Blockchain Analysis Report</title>
    <style>
        /* Base styles */
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, 'Open Sans', 'Helvetica Neue', sans-serif;
            margin: 0;
            padding: 0;
            color: #333;
            background-color: #f7f9fc;
            font-size: 14px;
        }}
        .container {{
            max-width: 100%;
            padding: 15px;
            margin: 0 auto;
        }}
        .report-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }}
        .report-title {{
            font-size: 20px;
            font-weight: bold;
            margin: 0;
            color: #2c3e50;
        }}
        .timestamp {{
            font-size: 12px;
            color: #7f8c8d;
        }}
        
        /* Tab styles */
        .tabs {{
            display: flex;
            flex-wrap: nowrap;
            overflow-x: auto;
            background-color: #fff;
            border-radius: 5px 5px 0 0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 0;
        }}
        .tab {{
            padding: 10px 15px;
            cursor: pointer;
            transition: background-color 0.3s;
            font-weight: 500;
            white-space: nowrap;
        }}
        .tab.active {{
            background-color: #3498db;
            color: white;
        }}
        .tab:hover:not(.active) {{
            background-color: #f1f1f1;
        }}
        
        /* Summary section styles */
        .summary-section {{
            background-color: white;
            border-radius: 0 0 5px 5px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            padding: 15px;
            margin-bottom: 15px;
        }}
        .wallet-address {{
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 15px;
            word-break: break-all;
            background-color: #f8f9fa;
            padding: 8px;
            border-radius: 4px;
            border-left: 4px solid #3498db;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
        }}
        .summary-card {{
            background-color: #f8f9fa;
            border-radius: 5px;
            padding: 12px;
        }}
        .summary-row {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 8px;
            border-bottom: 1px solid #eee;
            padding-bottom: 8px;
        }}
        .summary-row:last-child {{
            margin-bottom: 0;
            border-bottom: none;
            padding-bottom: 0;
        }}
        .summary-label {{
            font-weight: 500;
            color: #7f8c8d;
        }}
        .summary-value {{
            font-weight: 600;
            color: #2c3e50;
        }}
        
        /* Transaction table styles */
        .table-section {{
            background-color: white;
            border-radius: 5px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            padding: 15px;
            margin-bottom: 15px;
            overflow-x: auto;
        }}
        
        /* DataTables overrides for better alignment */
        table.dataTable {{
            width: 100% !important;
            margin: 0 !important;
        }}
        table.dataTable thead th, 
        table.dataTable tbody td {{
            padding: 8px 10px;
        }}
        table.dataTable thead th {{
            background-color: #f8f9fa;
            border-bottom: 2px solid #ddd;
        }}
        table.dataTable tbody td {{
            border-bottom: 1px solid #f2f2f2;
        }}
        .dataTables_wrapper .dataTables_filter input {{
            margin-left: 5px;
        }}
        
        /* Additional styles for specific elements */
        .positive {{
            color: #27ae60 !important;
        }}
        .negative {{
            color: #e74c3c !important;
        }}
        .monospace {{
            font-family: 'SFMono-Regular', Consolas, 'Liberation Mono', Menlo, monospace;
        }}
        .tab-content {{
            display: none;
        }}
        .tab-content.active {{
            display: block;
        }}
        
        /* Mobile optimizations */
        @media screen and (max-width: 768px) {{
            .container {{
                padding: 10px;
            }}
            .table-section {{
                padding: 10px;
            }}
            .dataTables_wrapper .dataTables_filter input {{
                width: 80px;
            }}
        }}
    </style>
    <!-- DataTables minimal CSS -->
    <link rel="stylesheet" href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css">
</head>
<body>
    <div class="container">
        <div class="report-header">
            <h1 class="report-title">Blockchain Transaction Analysis Report</h1>
            <div class="timestamp">Generated: {timestamp}</div>
        </div>
        
        <div class="tabs">
'''
    
    # Generate tabs for each sheet
    for i, sheet_name in enumerate(sheets_data.keys()):
        active_class = 'active' if i == 0 else ''
        html += f'            <div class="tab {active_class}" data-tab="{sheet_name}">{sheet_name}</div>\n'
    
    html += '''        </div>
        
'''
    
    # Generate content for each tab
    for i, (sheet_name, data) in enumerate(sheets_data.items()):
        active_class = 'active' if i == 0 else ''
        
        # Extract summary data
        summary = data['summary']
        address = summary['address']
        summary_rows = summary['summary_rows']
        
        # Extract main table data
        main_table = data['main_table']
        headers = main_table['headers']
        rows = main_table['rows']
        
        html += f'''        <div id="{sheet_name}-content" class="tab-content {active_class}">
            <div class="summary-section">
                <div class="wallet-address">{address}</div>
                <div class="summary-grid">
                    <div class="summary-card">
'''
        
        # Add first 2 summary rows
        for i, row in enumerate(summary_rows[:2]):
            label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
            value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
            
            # For start/end blocks, add monospace class
            if row['label'] in ['Start Block:', 'End Block:']:
                value_style += ' font-family: monospace;'
            
            html += f'''                        <div class="summary-row">
                            <span class="summary-label" style="{label_style}">{row['label']}</span>
                            <span class="summary-value" style="{value_style}">{row['value']}</span>
                        </div>
'''
        
        html += '''                    </div>
                    <div class="summary-card">
'''
        
        # Add last 2 summary rows
        for i, row in enumerate(summary_rows[2:]):
            label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
            value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
            html += f'''                        <div class="summary-row">
                            <span class="summary-label" style="{label_style}">{row['label']}</span>
                            <span class="summary-value" style="{value_style}">{row['value']}</span>
                        </div>
'''
        
        html += '''                    </div>
                    <div class="summary-card">
'''
        
        # Add extra summary info (from columns G and H)
        for i, row in enumerate(summary_rows):
            label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
            value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
            html += f'''                        <div class="summary-row">
                            <span class="summary-label" style="{label_style}">{row['extra_label']}</span>
                            <span class="summary-value" style="{value_style}">{row['extra_value']}</span>
                        </div>
'''
        
        html += '''                    </div>
                </div>
            </div>
            
            <div class="table-section">
                <table id="''' + f"{sheet_name}-table" + '''" class="display nowrap compact stripe">
                    <thead>
                        <tr>
'''
        
        # Add table headers
        for header in headers:
            header_style = ' '.join([f'{k}: {v};' for k, v in header['style'].items()])
            # For block #, ensure monospace formatting
            header_classes = 'monospace' if header['text'] == 'block #' else ''
            html += f'''                            <th style="{header_style}" class="{header_classes}">{header['text']}</th>
'''
        
        html += '''                        </tr>
                    </thead>
                    <tbody>
'''
        
        # Add table rows
        for row in rows:
            html += '''                        <tr>
'''
            for cell in row:
                cell_style = ' '.join([f'{k}: {v};' for k, v in cell['style'].items()])
                cell_text = cell['text']
                
                # Combine classes
                cell_classes = cell.get('class', '').strip()
                
                # Handle hyperlinks
                if cell['hyperlink']:
                    # For tx hashes, use the full text in the hyperlink but display shortened
                    display_text = cell['full_text'] if 'full_text' in cell else cell_text
                    html += f'''                            <td style="{cell_style}" class="{cell_classes}"><a href="{cell['hyperlink']}" target="_blank">{display_text}</a></td>
'''
                else:
                    html += f'''                            <td style="{cell_style}" class="{cell_classes}">{cell_text}</td>
'''
            
            html += '''                        </tr>
'''
        
        html += '''                    </tbody>
                </table>
            </div>
        </div>
'''
    
    # Add minimal JavaScript for interactivity
    html += '''        
    </div>

    <!-- Minimal required JavaScript -->
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
    
    <script>
        $(document).ready(function() {
            // Tab switching functionality
            $('.tab').on('click', function() {
                const tabId = $(this).attr('data-tab');
                
                // Update active tab
                $('.tab').removeClass('active');
                $(this).addClass('active');
                
                // Show active content
                $('.tab-content').removeClass('active');
                $('#' + tabId + '-content').addClass('active');
                
                // Adjust DataTable columns
                $('#' + tabId + '-table').DataTable().columns.adjust();
            });
            
            // Initialize DataTables with minimal options
            $('.tab-content table').each(function() {
                $(this).DataTable({
                    paging: true,
                    searching: true,
                    ordering: true,
                    pageLength: 25,
                    lengthMenu: [10, 25, 50, 100],
                    order: [[6, 'desc']], // Sort by Profit column (adjust index based on your columns)
                    scrollX: true,
                    deferRender: true,
                    language: {
                        search: "Filter:",
                        lengthMenu: "Show _MENU_",
                        info: "_START_ to _END_ of _TOTAL_"
                    },
                    dom: '<"top"lf>rt<"bottom"ip>',
                    columnDefs: [
                        // Set column-specific rendering options
                        {targets: 0, className: 'monospace'}, // Block number
                        {targets: 1, className: 'monospace'}, // TX hash
                    ]
                });
            });
        });
    </script>
</body>
</html>'''
    
    return html


# Example usage
if __name__ == "__main__":
    # Generate HTML report from Excel file
    excel_file = "MEV_daily_multiple_addresses_22016035_P&L.xlsx"
    html_output = "reports/blockchain_analysis_report.html"
    
    # Only keep these columns in the specified order
    columns_to_keep = [
        'block #', 'tx Hash', 'tx Index', 'miner',
        'Balance', 'Total Tip', 'Extractable Value', 'Profit',
        'Bribe %', 'Profit Rank', 'Resim', 'Resim @ txIx 0', 'Resim @ txIx -1'
    ]
    
    # Limit rows for size control (adjust as needed)
    max_rows_per_sheet = 100
    
    generate_html_report(excel_file, html_output, columns_to_keep, max_rows_per_sheet)
    print(f"HTML report generated: {html_output}")