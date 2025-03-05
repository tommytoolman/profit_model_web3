import os, time, json, math, requests
import backoff, logging, tempfile
import aiohttp, asyncio, nest_asyncio, subprocess
import pandas as pd
import numpy as np

from web3 import Web3
from decimal import Decimal
from eth_typing import HexStr
from hexbytes import HexBytes
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM
from datetime import datetime, timezone
from typing import Dict, List, Optional
from tqdm.notebook import tqdm
from pathlib import Path
from PIL import Image

from dotenv import load_dotenv
from lib_etherscan_funcs import get_address_tx_hashes_and_blocks
from chain_lib import w3_deejmon_http, deejmon_http, chain_data, minimal_abi
from block_calculation_engine import get_timestamps, fetch_block_number, main
from metadata import miner

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
# from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment

W3 = w3_deejmon_http

load_dotenv()

def connect_with_retries(web3, retries=5, delay=0.1):
    for attempt in range(retries):
        if web3.is_connected():
            print("Connected to Ethereum node!")
            return True
        else:
            print(f"Connection attempt {attempt + 1} failed. Retrying in {delay} seconds...")
            time.sleep(delay)
    return False

def restructure_df(df):

    restructured_data = {
        'block_number': {},
        'datetime': {},
        'timestamp': {},
    }

    # For each time period (current_datetime, midnight, etc.)
    for period in df.columns:
        # Get the values for this period
        values = df[period].iloc[0]  # Since we only have one row (Ethereum)

        # Add values to our restructured data
        restructured_data['timestamp'][period] = values['timestamp']
        restructured_data['datetime'][period] = values['datetime']
        restructured_data['block_number'][period] = values['block_number']

    return pd.DataFrame(restructured_data).transpose()

selected_periods = [ "current_datetime", "yesterday", "midnight", "yesterday_midnight", "start_of_week", "start_of_month"] # "start_of_1w"]

time_data = main(["Ethereum"], selected_periods)

address_list = os.environ.get("address_list").split(",")

start_block = int(time_data['yesterday']['Ethereum']['block_number'])

usdc_address = chain_data['Ethereum']['usdc_address']
usdc_proxy = Web3.to_checksum_address("0x43506849d7c04f9138d1a2050bbf3a0c054402dd")
usdt_address = chain_data['Ethereum']['usdt_address']
weth_address = chain_data['Ethereum']['weth_address']
stETH_address = "0xae7ab96520DE3A18E5e111B5EaAb095312D7fE84"

weth_contract = W3.eth.contract(address=weth_address, abi=minimal_abi)
usdt_contract = W3.eth.contract(address=usdt_address, abi=minimal_abi)
usdc_contract = W3.eth.contract(address=usdc_address, abi=minimal_abi)
stETH_contract = W3.eth.contract(address=stETH_address, abi=minimal_abi)

etherscan_api_key = os.environ.get("etherscan_api_key")

miner_map = {
    "0x95222290DD7278Aa3Ddd389Cc1E1d165CC4BAfe5" : "beaverbuild",
    "0x4838B106FCe9647Bdf1E7877BF73cE8B0BAD5f97" : "Titan Builder",
    "0x1f9090aaE28b8a3dCeaDf281B0F12828e676c326" : "rsync-builder.eth",
    "0x77777A6C097a1cE65C61A96a49bd1100F660eC94" : "MEV Builder: 0x777...C",
    "0x965Df5Ff6116C395187E288e5C87fb96CfB8141c" : "bloXroute: Builder 1",
    "0x388C818CA8B9251b393131C08a736A67ccB19297" : "Lido: Execution Layer Rewards Vault",
    "0xdadB0d80178819F2319190D340ce9A924f783711" : "BuilderNet",
    "0x7e2a2FA2a064F693f0a55C5639476d913Ff12D05" : "MEV Builder: 0x7e2...D05",
    "0xd4E96eF8eee8678dBFf4d535E033Ed1a4F7605b7" : "Rocket Pool Smoothing Pool",
    "0xe688b84b23f322a994A53dbF8E15FA82CDB71127" : "Fee Recipient: 0xe68...127",
    "0xd11D7D2cb0aFF72A61Df37fD016EE1bd9F180633" : "MEV Builder: 0xd11...633",
    "0x9f4Cf329f4cF376B7ADED854D6054859dd102a2A" : "Fee Recipient: 0x9f4...a2A",
}

token_contracts = {
    'WETH': weth_contract,
    'USDC': usdc_contract,
    'USDT': usdt_contract,
    'stETH': stETH_contract,
    }


def get_abi_from_etherscan(contract_address, etherscan_api_key):
    url = f"https://api.etherscan.io/api?module=contract&action=getabi&address={contract_address}&apikey={etherscan_api_key}"
    # print(url)
    response = requests.get(url)
    data = response.json()
    if data["status"] == "1":
        return data["result"]
    return None

def combine_token_balances(address, block_list, eth_results, weth_results, erc20_results_dict):
    """
    Combine ETH, WETH, and other ERC20 token balances into a structured format
    Args:
        block_list: List of block numbers
        eth_results: List of ETH balances from batch request
        weth_results: List of WETH balances from batch request
        erc20_results_dict: Dictionary mapping token symbols to lists of balances
    Returns:
        list: List of dictionaries containing block numbers and balances
    """
    # Ensure all lists are the same length
    if not (len(block_list) == len(eth_results) == len(weth_results)):
        raise ValueError("Mismatched lengths between blocks and results")

    # Also verify that each token in erc20_results_dict has the same length
    for token, balances in erc20_results_dict.items():
        if len(balances) != len(block_list):
            raise ValueError(f"Mismatched lengths for token {token}")

    combined_balances = []

    # Create a list of token names for consistent ordering
    token_names = sorted(erc20_results_dict.keys())

    # Zip all lists together and create dictionaries
    for i, (block, eth_bal, weth_bal) in enumerate(zip(block_list, eth_results, weth_results)):
        balance_dict = {
            "wallet_address": address,
            "block_number": int(block),
            "ETH": int(eth_bal),
            "WETH": int(weth_bal)
        }

        # Add other ERC20 token balances
        for token in token_names:
            balance_dict[token] = int(erc20_results_dict[token][i])

        combined_balances.append(balance_dict)

    return combined_balances

def get_token_balances_with_retry(W3, address, block_list, token_contracts, max_retries=3, batch_size=2500):
    """Get token balances with retry logic"""

    retry_count = 0
    while retry_count < max_retries:
        try:
            eth_results = []
            weth_results = []
            erc20_results_dict = {token: [] for token in token_contracts if token != 'WETH'}  # initialize the dictionary with empty lists for each token

            for i in range(0, len(block_list), batch_size):
                batch_blocks = block_list[i:i + batch_size]

                # Get ETH balances
                eth_batch = W3.batch_requests()
                for block in batch_blocks:
                    eth_batch.add(W3.eth.get_balance(address, block_identifier=int(block)))
                eth_results.extend(eth_batch.execute())

                # Get WETH balances
                weth_batch = W3.batch_requests()
                for block in batch_blocks:
                    weth_batch.add(token_contracts['WETH'].functions.balanceOf(address).call(block_identifier=int(block)))
                weth_results.extend(weth_batch.execute())

                for token, contract in token_contracts.items():
                    if token == 'WETH':
                        continue  # Skip WETH (already processed)

                    erc20_batch = W3.batch_requests()
                    for block in batch_blocks:
                        erc20_batch.add(contract.functions.balanceOf(address).call(block_identifier=int(block)))

                    # Execute the batch and append to the token's list
                    token_balances = erc20_batch.execute()
                    # print(token, token_balances)
                    erc20_results_dict[token].extend(token_balances)

                print(f"✅ Processed balances for blocks {i} to {i + len(batch_blocks)}")

            # print("\n")
            return combine_token_balances(address, block_list, eth_results, weth_results, erc20_results_dict)

        except Exception as e:
            retry_count += 1
            print(e)
            print(f"Balance retrieval attempt {retry_count} failed: {str(e)}")
            if retry_count < max_retries:
                print("Waiting before retry...")
                time.sleep(5)
            else:
                print("Max retries reached for balance retrieval\n")
                return []

def merge_on_different_keys(df1, df2, left_key, right_key, how='inner'):
    """
    Merges two DataFrames on columns with different names.

    Parameters:
        df1 (pd.DataFrame): The left DataFrame.
        df2 (pd.DataFrame): The right DataFrame.
        left_key (str): The key column in the left DataFrame.
        right_key (str): The key column in the right DataFrame.
        how (str): Type of merge (default is 'inner').

    Returns:
        pd.DataFrame: The merged DataFrame.
    """

    df1.loc[:, 'blockNumber'] = df1['blockNumber'].astype(int)
    df2.loc[:, 'block_number'] = df2['block_number'].astype(int)

    merged_df = pd.merge(df1, df2, left_on=left_key, right_on=right_key, how=how)
    return merged_df

def block_list_generator(block_results):
    block_details = []
    for count, n in enumerate(block_results):
        block_dict = {}
        for block in n:
            if block  in ["hash", "miner", "number", "timestamp", "baseFeePerGas"]:
                value = n[block]
                if isinstance(value, HexBytes):
                    value = f"0x{value.hex()}"
                block_dict[block] = value
        block_details.append(block_dict)
    return block_details

def tx_list_generator(tx_results):
    tx_details = []
    for count, n in enumerate(tx_results):
        tx_dict = {}
        for tx in n:
            if tx in ["type", "nonce", "maxFeePerGas", "maxPriorityFeePerGas", "baseFeePerGas", 'to', 'from', 'value', 'hash', "blockNumber", "transactionIndex", "sim_result_orig", "sim_result_0","sim_result_n1"]:
                value = n[tx]
                if isinstance(value, HexBytes):
                    value = f"0x{value.hex()}"
                tx_dict[tx] = value
        tx_details.append(tx_dict)
    return tx_details

def receipt_list_generator(receipt_results):
    receipts_list = []
    for count, n in enumerate(receipt_results):
        receipt_dict = {}
        for receipt in n:
            if receipt not in ["logs", "logsBloom"]:
                value = n[receipt]
                if isinstance(value, HexBytes):
                    value = f"0x{value.hex()}"
                receipt_dict[receipt] = value
        receipts_list.append(receipt_dict)
    return receipts_list

def trace_list_generator(block_details, trace_results, block_miners):
    """  TRACES STRUCTURE
        type(trace_results) = list
        type(trace_results[0]) = list
        type(trace_results[0][0]) = web3.datastructures.AttributeDict
        type(trace_results[0][0].action) = web3.datastructures.AttributeDict
        type(trace_results[0][0].action['from']) = str
    """
    trace_list = []
    for count, n in enumerate(trace_results):
        trace_dict = {}
        found_payment = False

        for trace in n:
            try:
                if trace.action['to'] == block_miners[trace.blockNumber]:
                    found_payment = True
                    trace_dict = {
                        "tx_hash": f"0x{trace.transactionHash.hex()}",
                        "from": trace.action['from'],
                        "to": trace.action['to'],
                        "payment_to_miner": f"{float(Web3.from_wei(trace.action['value'],'ether')):.6f}",
                    }
                    break
            except KeyError:
                trace_dict = {
                   "tx_hash": f"0x{trace.transactionHash.hex()}",
                   "from": None,
                   "to": None,
                   "payment_to_miner": None
                }

        if not found_payment and not trace_dict:
            trace_dict = {
               "tx_hash": f"0x{trace.transactionHash.hex()}",
               "from": None,
               "to": None, 
               "payment_to_miner": 0
            }

        trace_list.append(trace_dict)
    return trace_list

def process_initial_dataframe(df, address, balances):

    balances_df = {}

    balances_df = pd.DataFrame(balances)

    merged_df = merge_on_different_keys(df, balances_df, left_key='blockNumber', right_key='block_number', how='inner')

    # Drop the redundant column (optional)
    merged_df = merged_df.drop(columns=['block_number'])
    merged_df = merged_df.sort_values(by="blockNumber")
    sorted_df = merged_df.sort_values(by="blockNumber", ignore_index=True)

    # Convert 'wei' column to ether and round to n decimal places and sum Totals
    sorted_df['ETH'] = sorted_df['ETH'].apply(lambda x: round(Web3.from_wei(x, 'ether'), 6))
    sorted_df['WETH'] = sorted_df['WETH'].apply(lambda x: round(Web3.from_wei(x, 'ether'), 6))
    sorted_df['Total'] = sorted_df['ETH'] + sorted_df['WETH']
    weth_index = sorted_df.columns.get_loc('WETH')
    sorted_df.insert(weth_index + 1, 'Total', sorted_df.pop('Total'))

    try:
        sorted_df['value'] = sorted_df['value'].apply(
            lambda x: '{:.0f}'.format(0) if float(x) == 0
            else '{:.3f}'.format(float(Web3.from_wei(int(x), 'ether')))
        )
    except:
        pass

    # Ensure the timestamp is an integer
    sorted_df['timeStamp'] = sorted_df['timeStamp'].astype(int)
    sorted_df['dateTime'] = sorted_df['timeStamp'].apply(lambda x: datetime.fromtimestamp(x, tz=timezone.utc).strftime('%d-%m-%Y %H:%M:%S'))
    ts_index = sorted_df.columns.get_loc('timeStamp')
    sorted_df.insert(ts_index + 1, 'dateTime', sorted_df.pop('dateTime'))

    # Step 1: Preprocess Data to Handle Missing Values
    # First, ensure methodId uses empty strings instead of NaN, and nonce uses 0 instead of NaN where appropriate:
    try:
        sorted_df['methodId'] = sorted_df['methodId'].fillna('')
    except:
        print("No methodID!")
    # Convert nonce to Int64 (handles NaN gracefully)
    sorted_df['nonce'] = sorted_df['nonce'].astype('Int64')
    # Replace NaN in methodId with empty strings
    sorted_df['methodId'] = sorted_df['methodId'].fillna('')

    # Step 2: Define Robust Aggregation Logic
    # Use these functions to handle edge cases:
    def get_valid_nonce(series):
        """Return first non-zero, non-NaN value. Fallback to 0 if none exist."""
        valid_values = series[series != 0].dropna()
        return valid_values.iloc[0] if not valid_values.empty else 0

    def get_valid_method_id(series):
        """Return first non-blank value. Fallback to '' if none exist."""
        valid_values = series[series != ''].dropna()
        return valid_values.iloc[0] if not valid_values.empty else ''

    def combine_token_types(series):
        """Merge unique tokenType values into a sorted list."""
        unique_values = series.explode().dropna().unique()
        return sorted(unique_values) if len(unique_values) > 1 else unique_values[0]


    def identify_withdrawals(row):
        if (row['methodId'].startswith('0x') and 
            pd.isna(row['functionName']) and 
            row['from'].lower() == address.lower() and 
            float(row['Profit']) < 0):
            return 'Withdrawal'
        return row['functionName']

    sorted_df['functionName'] = sorted_df.apply(identify_withdrawals, axis=1)

    # Step 3: Apply Aggregation
    aggregation_rules = {
        'tokenType': combine_token_types,
        'nonce': get_valid_nonce,
        'methodId': get_valid_method_id,
        'functionName': 'first',
        'dateTime': 'first',
        'from': 'first',
        'to': 'first',
        'value': 'first',
        'ETH': 'first',
        'WETH': 'first',
        'Total': 'first',
        'USDC': 'first',
        'USDT': 'first',
        'stETH': 'first',
    }

    # Group by blockNumber and txHash, then aggregate
    sorted_df = sorted_df.groupby(['blockNumber', 'txHash'], as_index=False).agg(aggregation_rules)

    # Calculate profit/loss as the difference between current Total and previous Total
    sorted_df['Profit'] = (sorted_df['Total'] - sorted_df['Total'].shift(1)).round(4)

    # Skip the first row (no previous row to compare)
    sorted_df.loc[0, 'Profit'] = None

    # Rank profits in descending order (highest profit = rank 1)
    sorted_df['Profit_Rank'] = sorted_df['Profit'].rank(ascending=False, method='min').astype('Int64')

    # Handle NaN values (e.g., first row) by assigning the lowest rank
    max_rank = sorted_df['Profit_Rank'].max()
    sorted_df['Profit_Rank'] = sorted_df['Profit_Rank'].fillna(max_rank + 1).astype('Int64')

    final_df = sorted_df.sort_values(by="Profit_Rank")

    final_df['from'] = final_df['from'].apply(lambda x: x[0:8])
    final_df['to'] = final_df['to'].apply(lambda x: x[0:8])

    final_df['ETH'] = final_df['ETH'].apply(lambda x: round(x, 2))
    final_df['WETH'] = final_df['WETH'].apply(lambda x: round(x, 2))
    final_df['Total'] = final_df['Total'].apply(lambda x: round(x, 2))

    return final_df

def combine_blockchain_data(block_details, tx_details, receipt_details, trace_details, balances):

    # Create mappings for faster lookups
    block_map = {b['number']: b for b in block_details}

    # Add defensive receipt mapping
    receipt_map = {}
    for r in receipt_details:
        try:
            receipt_map[r['transactionHash']] = {
                'gasUsed': r.get('gasUsed', 0),
                'effectiveGasPrice': r.get('effectiveGasPrice', 0)
            }
        except KeyError as e:
            print(f"Warning: Missing receipt data: {str(e)}")
            receipt_map[r['transactionHash']] = {
                'gasUsed': 0,
                'effectiveGasPrice': 0
            }
    
    trace_map = {t.get('tx_hash'): t for t in trace_details if t.get('tx_hash')}
    balance_map = {b['wallet_address']: {k:v for k,v in b.items() if k != 'block_number'} 
                  for b in balances}

    combined_list = []

    for tx in tx_details:
        combined_dict = tx.copy()

        # Add block data
        if tx['blockNumber'] in block_map:
            block = block_map[tx['blockNumber']]
            combined_dict.update({
                'miner': block.get('miner', ''),
                'timestamp': block.get('timestamp', 0),
                'baseFeePerGas': block.get('baseFeePerGas', 0)
            })

        # Add receipt data
        if tx['hash'] in receipt_map:
            combined_dict.update(receipt_map[tx['hash']])
        else:
            combined_dict.update({
                'gasUsed': 0,
                'effectiveGasPrice': 0
            })

        # Add trace data
        if tx['hash'] in trace_map:
            combined_dict['payment_to_miner'] = trace_map[tx['hash']].get('payment_to_miner', 0)
        else:
            combined_dict['payment_to_miner'] = 0

        # Add balance data
        if tx['from'] in balance_map:
            combined_dict['balances'] = balance_map[tx['from']]
        elif tx['to'] in balance_map:
            combined_dict['balances'] = balance_map[tx['to']]

        combined_list.append(combined_dict)

    # Defensive gas calculations
    for tx in combined_list:
        try:
            tx['base_gas_fee'] = tx.get('gasUsed', 0) * tx.get('baseFeePerGas', 0)
            tx['base_gas_fee_eth'] = round(Web3.from_wei(tx.get('gasUsed', 0) * tx.get('baseFeePerGas', 0), 'ether'), 4)
            tx['gas_cost'] = tx.get('gasUsed', 0) * tx.get('effectiveGasPrice', 0)
            tx['gas_cost_eth'] = round(Web3.from_wei(tx.get('gas_cost', 0), 'ether'), 4)
            tx['max_possible_priority_fee'] = max(0, tx.get('maxFeePerGas', 0) - tx.get('baseFeePerGas', 0))
            tx['effective_priority_fee'] = min(tx.get('maxPriorityFeePerGas', 0), tx.get('max_possible_priority_fee', 0))
            tx['gas_tip'] = tx.get('gasUsed', 0) * tx.get('effective_priority_fee', 0)
            tx['total_gas_fee'] = tx.get('base_gas_fee', 0) + tx.get('gas_tip', 0)
            tx['gas_fee_bribe'] = round(float(Web3.from_wei(tx.get('gas_tip', 0), 'ether')), 4)
            tx['direct_miner_bribe'] = round(float(tx.get('payment_to_miner', 0)) if tx.get('payment_to_miner') else 0.0, 4)
            tx['total_tip'] = tx.get('gas_fee_bribe', 0) + tx.get('direct_miner_bribe', 0)
            tx['total_tx_cost'] = tx.get('total_gas_fee', 0) + tx.get('direct_miner_bribe', 0)
            tx['total_tx_cost_eth'] = round(Web3.from_wei(tx.get('total_tx_cost', 0), 'ether'), 4)
        except Exception as e:
            print(f"Warning: Error calculating gas fees for tx {tx.get('hash', 'unknown')}: {str(e)}")
            # Set default values for all calculated fields
            default_fields = [
                'base_gas_fee', 'base_gas_fee_eth', 'gas_cost', 'gas_cost_eth',
                'max_possible_priority_fee', 'effective_priority_fee', 'gas_tip',
                'total_gas_fee', 'gas_fee_bribe', 'direct_miner_bribe',
                'total_tip', 'total_tx_cost', 'total_tx_cost_eth'
            ]
            for field in default_fields:
                tx[field] = 0

    return combined_list

def merge_dataframes(initial_df, combined_df, libmev_df):

    merged_df = initial_df.merge(
        combined_df[['maxPriorityFeePerGas', 'maxFeePerGas', 'transactionIndex', 'miner', 'baseFeePerGas', 'gasUsed', 'effectiveGasPrice', 'payment_to_miner', 'balances', 'base_gas_fee',
                     'base_gas_fee_eth', 'gas_cost', 'gas_cost_eth', 'max_possible_priority_fee', 'effective_priority_fee', 'gas_tip', 'total_gas_fee', 'gas_fee_bribe', 'direct_miner_bribe',
                     'total_tip', 'total_tx_cost', 'total_tx_cost_eth',  'sim_result_orig', 'sim_result_0', 'sim_result_n1', 'hash']],
        left_on='txHash',
        right_on='hash',
        how='left',
        validate='1:1'  # Ensure 1-to-1 merge
    )

    merged_df['extractable_value'] = (
        round(pd.to_numeric(merged_df['Profit']) + 
        merged_df['total_tip'] +
        merged_df['base_gas_fee_eth'].astype(float), 4)
    )

    merged_df['USDC'] = merged_df['USDC'].apply(lambda x: f"${(x/1000000/1000000):,.3f}m")
    merged_df['USDT'] = merged_df['USDT'].apply(lambda x: f"${(x/1000000/1000000):,.3f}m")

    merged_df['Profit'] = round(merged_df['Profit'].astype(float), 4)
    merged_df['extractable_value'] = merged_df['extractable_value'].astype(float)

    merged_df['profit_margin'] = np.where(
       merged_df['extractable_value'] != 0,
       (merged_df['Profit'] / merged_df['extractable_value'] * 100).round(1).astype(str) + '%',
       '0%'
    )

    merged_df.insert(
        loc=merged_df.columns.get_loc('miner') + 1,
        column='miner_name',
        value=merged_df['miner'].apply(lambda x: miner_map.get(x, 'Other'))
    )

    merged_df = merged_df[[
            'blockNumber', 'dateTime', 'txHash', 'transactionIndex', 'nonce', 'miner_name', 'Total', 'methodId', 'functionName',
            # 'USDT', 'USDC',
            'base_gas_fee_eth', 'gas_fee_bribe', 'direct_miner_bribe', 'total_tip', 'total_tx_cost_eth', 'extractable_value',
            'Profit', 'profit_margin', 'Profit_Rank', 'sim_result_orig', 'sim_result_0', 'sim_result_n1',
            ]]

    merged_df = merged_df.rename(columns={
        'blockNumber' : 'block #',
        'txHash': 'tx Hash',
        'miner_name' : 'miner',
        'Total': 'Balance (ETH + WETH)',
        'functionName' : 'Function',
        'transactionIndex': 'tx Index',
        'gas_fee_bribe': 'Gas Fee Bribe',
        'base_gas_fee_eth': 'Base Gas Fee (ETH)',
        'direct_miner_bribe': 'Miner Bribe',
        'total_tip' : 'Total Tip',
        'total_tx_cost_eth' : 'Tx Cost (ETH)',
        'extractable_value' : 'Extractable Value',
        # 'gas_fee_bribe', 'direct_miner_bribe', 'total_tip', 'total_tx_cost_eth', 'extractable_value',
        'profit_margin': 'Margin %',
        'Profit_Rank' : 'Profit Rank',
        'sim_result_orig': 'Resim',
        'sim_result_0': 'Resim @ txIx 0',
        'sim_result_n1': 'Resim @ txIx -1'
    })

    return merged_df

async def resim_tx_async(session, deejmon_http, resim_bundle, tx_index):
    payload = {
        "id": 1,
        "jsonrpc": "2.0",
        "method": "debug_traceCallMany",
        "params": [
            [
                {
                    "blockOverride": {},
                    "transactions": [
                        {
                            "from": resim_bundle['from'],
                            "to": resim_bundle['to'],
                            "data": resim_bundle['data'],
                            "gas": resim_bundle['gas'],
                            "value": resim_bundle['value'],
                        }
                    ]
                }
            ],
            {
                "blockNumber": hex(resim_bundle['blockNumber']),
                "transactionIndex": tx_index,
            },
            {
                "tracer": "callTracer",
                "tracerConfig": {
                    "diffMode": False,
                    "onlyTopCall": True,
                    "withLog": False
                }
            }
        ]
    }
    headers = {
        "accept": "application/json",
        "content-type": "application/json"
    }

    async with session.post(deejmon_http, json=payload, headers=headers) as response:
        result = await response.json()
        return {
            'tx_index': tx_index,
            'resim_bundle': resim_bundle,
            'result': result
        }

async def process_transactions(tx_results_dicts, deejmon_http):
    async with aiohttp.ClientSession() as session:
        tasks = []
        for count, tx in enumerate(tx_results_dicts):
            resim_bundle = {
                "from": tx['from'],
                "to": tx['to'],
                "data": tx['input'].hex() if hasattr(tx['input'], 'hex') else tx['input'],
                "gas": tx['gas'],
                "value": tx['value'],
                "blockNumber": tx['blockNumber'],
            }

            # Create tasks in order but don't await them yet
            task0 = resim_tx_async(session, deejmon_http, resim_bundle, tx['transactionIndex'])
            task1 = resim_tx_async(session, deejmon_http, resim_bundle, 0)
            task2 = resim_tx_async(session, deejmon_http, resim_bundle, -1)
            tasks.append((count, [task0, task1, task2]))  # Keep track of original order

        # Use gather to maintain order while still running async
        results = []
        for count, task_group in tqdm(tasks):
            group_results = await asyncio.gather(*task_group)
            results.extend(group_results)

        return results

async def main(tx_results):
    if not connect_with_retries(W3, retries=3, delay=0.001):
        raise Exception("Failed to connect to the Ethereum node after multiple attempts")

    tx_results_dicts = [dict(tx) for tx in tx_results]
    
    print(f"Processing {len(tx_results_dicts)} transactions...")
    results = await process_transactions(tx_results_dicts, deejmon_http)
    
    # Process results in guaranteed order
    for i in range(0, len(results), 3):
        tx_index = i // 3
        result_orig = results[i]      # Original tx index result
        result_0 = results[i + 1]     # Index 0 result
        result_n1 = results[i + 2]    # Index -1 result

        tx_results_dicts[tx_index]['sim_result_orig'] = parse_sim_result(result_orig['result'])
        tx_results_dicts[tx_index]['sim_result_0'] = parse_sim_result(result_0['result'])
        tx_results_dicts[tx_index]['sim_result_n1'] = parse_sim_result(result_n1['result'])

        block_num = tx_results_dicts[tx_index]['blockNumber']
        # print(f"Block {block_num} || {tx_results_dicts[tx_index]['sim_result_orig']} || {tx_results_dicts[tx_index]['sim_result_0']} || {tx_results_dicts[tx_index]['sim_result_n1']}")

    return tx_results_dicts

def parse_sim_result(result):
    if "error" in result:
        return result
    else:
        if "error" in result['result'][0][0]:
            try:
                return f"Revert:{result['result'][0][0]['revertReason']}"
            except KeyError:
                return f"Error:{result['result'][0][0]['error']}"
        else:
            return "OK"

def create_final_df(libmev_df, merged_df):

    try:

        exploded_df = libmev_df.explode('txs')

        # If you want to check if the tx Hash is in any of the arrays in txs
        merged_df['matches'] = merged_df['tx Hash'].apply(lambda x: any(x in val for val in libmev_df['txs']))
    
        # First merge with exploded data
        merged_result = merged_df.merge(
            exploded_df,
            left_on='tx Hash',
            right_on='txs',
            how='left',
            suffixes=('', '_exploded')
        )
    
        # Then merge with original array data
        final_result = merged_result.merge(
            libmev_df,
            on='bundle_hash',  # assuming bundle_hash is the key in libMEV_df
            how='left',
            suffixes=('', '_original')
        )
    
        # print(final_result.keys())
        
        final_result = final_result[[
            'block #', 'dateTime', 'tx Hash', 'tx Index', 'tags', 'miner',
           'Balance (ETH + WETH)', 'methodId', 'Function',
           'Base Gas Fee (ETH)', 'Gas Fee Bribe', 'Miner Bribe',
           'Total Tip', 'Tx Cost (ETH)', 'Extractable Value', 'Profit',
           'Margin %', 'Profit Rank', 'Resim', 'Resim @ txIx 0',
           'Resim @ txIx -1',
            'txs_original', 'txs_count_original',
            'nonce', 'timestamp', 'bundle_hash', 
            'searcher_txs_original', 'searcher_txs_count_original',
           'searcher_eoa', 'searcher_contract', 'tokens', 'tokens_count',
           'token_balance_delta', 'builder_address', 'extra_data',
           'burned_eth', 'tipped_eth', 'profit_eth', 'profit_margin', 'burned_usdc', 'tipped_usdc', 'profit_usdc',
           # 'searcher_gas_used', 'bundle_gas_used', 'matches',
            ]]

        final_result = final_result.rename(columns={
            'tags' : 'MEV Type',
            'txs_original' : 'Txs in Bundle',
            'searcher_txs_original': 'searcher_txs',
            'searcher_txs_count_original' : 'searcher_txs_count',
        })

        try:
            final_result = final_result.rename(columns={
                'txs_count_original' : '# Txs in Bundle',
            })
        except:
            pass

        return final_result

    except KeyError:

        # Case when libmev does not have that contract logged so no data
        final_result = merged_df

        return final_result

class BundleFetcher:
   def __init__(self, web3_url: str, max_concurrent: int = 10, retry_attempts: int = 5):
       self.max_concurrent = max_concurrent
       self.retry_attempts = retry_attempts
       self.w3 = Web3(Web3.HTTPProvider(web3_url))
       self.pbar = None
       self.stats = {
           'requests_made': 0,
           'retries': 0,
           'failed_requests': 0,
           'total_data_points': 0,
           'start_time': None,
           'end_time': None
       }
       self.state_file = Path('bundle_fetch_state.json')
       logging.basicConfig(level=logging.INFO)
       self.logger = logging.getLogger(__name__)

   def load_state(self, address: str) -> Dict:
       if self.state_file.exists():
           with open(self.state_file, 'r') as f:
               return json.load(f).get(address, {})
       return {}

   def save_state(self, address: str, latest_block: int, count: int):
       state = {}
       if self.state_file.exists():
           with open(self.state_file, 'r') as f:
               state = json.load(f)

       state[address] = {
           'latest_block': latest_block,
           'count': count,
           'last_updated': datetime.now().isoformat(),
           'stats': self.stats
       }

       with open(self.state_file, 'w') as f:
           json.dump(state, f, indent=2)

   @backoff.on_exception(
       backoff.expo,
       (aiohttp.ClientError, asyncio.TimeoutError, ConnectionResetError),
       max_tries=5,
       max_time=300
   )
   async def fetch_with_delay(self, session: aiohttp.ClientSession, url: str) -> Dict:
       self.stats['requests_made'] += 1
       try:
           timeout = aiohttp.ClientTimeout(total=30)
           async with session.get(url, timeout=timeout) as response:
               if response.status != 200:
                   self.stats['failed_requests'] += 1
                   raise aiohttp.ClientError(f"API returned status {response.status}")
               data = await response.json()
               await asyncio.sleep(2)
               return data
       except Exception as e:
           self.stats['retries'] += 1
           raise

   async def fetch_all_bundle_data(self, address: str, from_block: Optional[int] = None, to_block: Optional[int] = None) -> List[Dict]:
       base_url = "https://api.libmev.com/v1/bundles"
       all_data = []
       failed_offsets = []

       latest_block = self.w3.eth.get_block("latest")['number']
       block_range = f"{from_block},{to_block}" if from_block else f"0,{latest_block}"

       self.stats['start_time'] = time.time()

       try:
           connector = aiohttp.TCPConnector(
               limit=self.max_concurrent,
               ttl_dns_cache=300,
               force_close=True
           )

           async with aiohttp.ClientSession(connector=connector) as session:
               url_params = f"blockRange={block_range}&filterByContracts={address}&limit=50&orderByAsc=block_number"
               initial_url = f"{base_url}?{url_params}&offset=0"
               initial_response = await self.fetch_with_delay(session, initial_url)

               total_count = initial_response['count']
               all_data.extend(initial_response['data'])

               self.pbar = tqdm(total=total_count, desc=f"Fetching libMEV bundles for {address[:10]}...")
               self.pbar.update(len(initial_response['data']))

               semaphore = asyncio.Semaphore(self.max_concurrent)
               offsets = range(50, total_count, 50)

               async def fetch_page(offset: int) -> List[Dict]:
                   async with semaphore:
                       url = f"{base_url}?{url_params}&offset={offset}"
                       try:
                           response = await self.fetch_with_delay(session, url)
                           if response.get('data'):
                               self.pbar.update(len(response['data']))
                               return response['data']
                           return []
                       except Exception as e:
                           failed_offsets.append(offset)
                           return []

               if offsets:
                   results = await asyncio.gather(*[fetch_page(offset) for offset in offsets])
                   for result in results:
                       all_data.extend(result)

               self.stats['end_time'] = time.time()
               self.stats['total_data_points'] = len(all_data)
               self.stats['duration'] = self.stats['end_time'] - self.stats['start_time']

               self.save_state(address, latest_block, len(all_data))

               if failed_offsets:
                   self.logger.warning(f"Failed offsets: {failed_offsets}")

               return all_data

       finally:
           if self.pbar:
               self.pbar.close()

def get_bundles(address: str, web3_url: str, from_block: Optional[int] = None,  to_block: Optional[int] = None, max_concurrent: int = 12) -> Dict:
   try:
       loop = asyncio.get_event_loop()
   except RuntimeError:
       loop = asyncio.new_event_loop()
       asyncio.set_event_loop(loop)

   fetcher = BundleFetcher(web3_url=web3_url, max_concurrent=max_concurrent)
   bundles = loop.run_until_complete(fetcher.fetch_all_bundle_data(address, from_block, to_block))

   return {
       'bundles': bundles,
       'stats': fetcher.stats,
       'status': 'completed',
       'address': address,
       'timestamp': datetime.now().isoformat()
   }

def libmev_data(combined_df):
    """
    We take combined_df, get all contracts called by address or addresses.
    Get min / max block numbers and then pass to libmev API.
    """

    unique_vals = combined_df['to'].unique().tolist()

    result = combined_df.groupby('to')['blockNumber'].agg(['min', 'max'])
    result_list_of_dicts = (
        result
        .reset_index()
        .rename(columns={'index': 'to'})
        .to_dict('records')
    )

    result_list = []

    for n in result_list_of_dicts:
        result = {}
        result = get_bundles(
            address=n['to'],
            web3_url=deejmon_http,
            from_block=n['min'],
            to_block=n['max']
        )

        result_list.extend(result['bundles'])

    print(f"Fetch took {result['stats']['duration']:.2f} seconds")
    print(f"Success rate: {(result['stats']['requests_made'] - result['stats']['failed_requests']) / result['stats']['requests_made'] * 100:.2f}%")

    bundles_df = pd.DataFrame(result_list)
    try:
        bundles_df.sort_values(by="block_number", ascending=True)
    except:
        return bundles_df

    def count_array(x):
        if isinstance(x, list):
            return len(x)
        elif isinstance(x, str):
            try:
                # If it's a string representation of an array
                return len(eval(x))
            except:
                return 1 if x else 0
        else:
            return 0

    array_col_list = ['tokens', 'txs', 'searcher_txs']

    for array_col in array_col_list:
        position = bundles_df.columns.get_loc(array_col) + 1
        bundles_df.insert(
            position,
            f'{array_col}_count',  # name of the new column
            bundles_df[array_col].apply(count_array)
        )

    return bundles_df

nest_asyncio.apply()

def process_address(address, start_block):
    """Process a single address"""
    address = Web3.to_checksum_address(address.lower())

    L24_tx = get_address_tx_hashes_and_blocks(address, start_block)
    df = pd.DataFrame(L24_tx)

    abi = get_abi_from_etherscan(usdc_proxy, etherscan_api_key)

    contract_new = W3.eth.contract(address=usdc_address, abi=abi)

    tx_list = list(set([x['txHash'] for x in L24_tx]))
    unique_blocks = list(set(df['blockNumber'].to_list()))
    latest_block = max(unique_blocks)

    # Get token balances
    balances = get_token_balances_with_retry(W3, address, unique_blocks, token_contracts)

    # Fetch block data
    batch = W3.batch_requests()
    for block in unique_blocks:
        batch.add(W3.eth.get_block(int(block)))
    block_results = batch.execute()
    block_miners = {block.number: block.miner for block in block_results}
    block_details = block_list_generator(block_results)
    print(f"\n✅ Fetched blocks")

    # Fetch transaction data
    batch = W3.batch_requests()
    for tx in tx_list:
        batch.add(W3.eth.get_transaction(tx))
    tx_results = batch.execute()
    print(f"✅ Fetched transactions")

    # Fetch receipt data
    batch = W3.batch_requests()
    for tx in tx_list:
        batch.add(W3.eth.get_transaction_receipt(tx))
    receipt_results = batch.execute()
    receipt_details = receipt_list_generator(receipt_results)
    print(f"✅ Fetched transaction receipts")

    # Fetch trace data
    batch = W3.batch_requests()
    for tx_hash in tx_list:
        tx_hash = tx_hash.hex() if isinstance(tx_hash, HexBytes) else tx_hash
        batch.add(W3.tracing.trace_transaction(tx_hash))
    trace_results = batch.execute()
    trace_details = trace_list_generator(block_results, trace_results, block_miners)
    print(f"✅ Fetched transaction traces\n")

    updated_tx_results = asyncio.run(main(tx_results))
    tx_details = tx_list_generator(updated_tx_results)

    # Process dataframes
    initial_df = process_initial_dataframe(df, address, balances)
    combined_data = combine_blockchain_data(block_details, tx_details, receipt_details, trace_details, balances)
    combined_df = pd.DataFrame(combined_data)
    libmev_df = libmev_data(combined_df)
    merged_df = merge_dataframes(initial_df, combined_df, libmev_df)
    final_df = create_final_df(libmev_df, merged_df)

    # print(final_df.keys())
    
    return combined_df, merged_df, libmev_df, final_df, latest_block

def format_pl_report(address, df, output_xlsx, wb=None, ws=None):
    """
    Format PL report with professional styling
    Args:
        address: wallet address
        df: dataframe to write
        output_xlsx: output filename
        wb: optional workbook object (for multiple addresses)
        ws: optional worksheet object (for multiple addresses)
    """

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.dimensions import ColumnDimension
    import math  # For isnan check

    def set_sheet_white(ws):
        """
        Set entire sheet to white background with no grid lines
        """
        from openpyxl.styles import PatternFill
        
        # Create white fill
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Apply to all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = white_fill
        
        # Turn off grid lines
        ws.sheet_view.showGridLines = False

    def convert_to_excel_friendly(value):
        if isinstance(value, list):
            return ', '.join(str(x) for x in value)
        elif isinstance(value, dict):
            if not value:
                return ''
            parts = []
            for k, v in value.items():
                if isinstance(v, dict):
                    nested_parts = [f"{sub_k}: {sub_v}" for sub_k, sub_v in v.items()]
                    parts.append(f"{k}: {{{', '.join(nested_parts)}}}")
                else:
                    parts.append(f"{k}: {v}")
            return '{' + ', '.join(parts) + '}'
        return value

    def clean_illegal_excel_chars(value):
        """Clean string values to remove illegal Excel characters"""
        if isinstance(value, str):
            # Handle common resim cases
            resim_replacements = {
                'Revertᴿ': 'Revert:R',
                'Revert:Rˡ': 'Revert:R',
                'ˡ': 'l',
                'ᴿ': 'R',
                '\u0000': '',  # Null character
            }

            for old, new in resim_replacements.items():
                value = value.replace(old, new)

            # Remove any other control characters
            value = ''.join(char for char in value if ord(char) >= 32 or char in '\n\r\t')

        return value

    def safe_numeric(value):
        """
        Safely convert any numeric type to float.
        """
        if value is None:
            # print(f"DEBUG safe_numeric: Got None value")
            return None

        # print(f"DEBUG safe_numeric: Converting value: {value} (type: {type(value)})")
        try:
            result = float(value)
            # print(f"DEBUG safe_numeric: Converted to: {result}")
            return result
        except (ValueError, TypeError) as e:
            # print(f"DEBUG safe_numeric: Conversion failed with error: {str(e)}")
            return None

    def get_column_letter_by_header(worksheet, header_text, header_row=9):
        """
        Find the column letter based on the header text in the specified row.
        
        Args:
            worksheet: The openpyxl worksheet
            header_text: The text to search for in the header row
            header_row: The row number containing headers (default 9)
        
        Returns:
            str: Column letter (e.g., 'A', 'B', 'AA') or None if not found
        """
        from openpyxl.utils import get_column_letter
        
        for cell in worksheet[header_row]:
            if cell.value == header_text:
                return get_column_letter(cell.column)
        return None

    # Clean the dataframe
    for column in df.columns:
        df[column] = df[column].apply(lambda x:
            f"Error: {x['error']['message']}" if isinstance(x, dict) and 'error' in x
            else convert_to_excel_friendly(x)
        )

    # Then modify the DataFrame cleaning section to use both functions:
    for column in df.columns:
        df[column] = df[column].apply(lambda x:
            clean_illegal_excel_chars(
                f"Error: {x['error']['message']}" if isinstance(x, dict) and 'error' in x
                else convert_to_excel_friendly(x)
            )
        )

    # Initial Excel setup
    if wb is None or ws is None:
        df.to_excel(output_xlsx, sheet_name='PL Report', startrow=8, startcol=1, index=False)
        wb = load_workbook(output_xlsx)
        ws = wb['PL Report']
        set_sheet_white(ws)
    else:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 9):
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx, column=c_idx, value=value)
        set_sheet_white(ws)


    # Handle exception for 0x75efe to replace my profit calcs with those from libmev - for now at least
    if address == "0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D":
        # print(f"HANDLING EXCEPTION FOR {address} ")

        # Get column letters based on headers
        profit_col = get_column_letter_by_header(ws, 'Profit')
        profit_eth_col = get_column_letter_by_header(ws, 'profit_eth')
        margin_col = get_column_letter_by_header(ws, 'Margin %')
        profit_margin_col = get_column_letter_by_header(ws, 'profit_margin')
        tx_cost_col = get_column_letter_by_header(ws, 'Tx Cost (ETH)')
        extractable_col = get_column_letter_by_header(ws, 'Extractable Value')
        rank_col = get_column_letter_by_header(ws, 'Profit Rank')

        # print(f"\nDEBUG: Found columns:")
        # print(f"profit_col: {profit_col}")
        # print(f"profit_eth_col: {profit_eth_col}")
        # print(f"margin_col: {margin_col}")
        # print(f"profit_margin_col: {profit_margin_col}")
        
        # Get the last row with data
        last_row = ws.max_row
        # print(f"DEBUG: Last row is {last_row}")
        
        # 1. Replace Profit values with profit_eth
        if profit_col and profit_eth_col:
            # print("\nDEBUG: Copying profit_eth values to profit column")
            for row in range(10, last_row + 1):
                profit_eth_cell = ws[f'{profit_eth_col}{row}']
                profit_cell = ws[f'{profit_col}{row}']
                
                # if row <= 15:  # Debug first few rows
                    # print(f"\nDEBUG Row {row}:")
                    # print(f"Original profit_eth value: {profit_eth_cell.value} ({type(profit_eth_cell.value)})")
                
                # Convert None, empty strings, or NaN to 0
                if profit_eth_cell.value in (None, '', 'nan') or (isinstance(profit_eth_cell.value, float) and math.isnan(profit_eth_cell.value)):
                    profit_value = 0.0
                else:
                    profit_value = safe_numeric(profit_eth_cell.value)
                    if profit_value is None:  # If conversion failed
                        profit_value = 0.0
                
                profit_cell.value = profit_value
                profit_cell.number_format = '0.0000'  # 4 decimal places
                
                # if row <= 15:
                #     print(f"Converted profit value: {profit_value} ({type(profit_value)})")

        # 2. Replace Margin % with profit_margin and format as percentage
        if margin_col and profit_margin_col:
            # print("\nDEBUG: Copying profit_margin values to margin column")
            for row in range(10, last_row + 1):
                margin_cell = ws[f'{margin_col}{row}']
                profit_margin_cell = ws[f'{profit_margin_col}{row}']
                margin_value = safe_numeric(profit_margin_cell.value)
                if margin_value is not None:
                    margin_cell.value = margin_value
                    margin_cell.number_format = '0.0%'  # Percentage with 1 decimal

        # 3. Recalculate Extractable Value as Profit + Tx Cost
        if tx_cost_col and profit_col and extractable_col:
            # print("\nDEBUG: Recalculating Extractable Value")
            for row in range(10, last_row + 1):
                tx_cost_cell = ws[f'{tx_cost_col}{row}']
                profit_cell = ws[f'{profit_col}{row}']
                extractable_cell = ws[f'{extractable_col}{row}']
                
                tx_cost = safe_numeric(tx_cost_cell.value)
                profit = safe_numeric(profit_cell.value)
                if tx_cost is not None and profit is not None:
                    extractable_cell.value = tx_cost + profit
                    extractable_cell.number_format = '0.0000'

        # 4. Sort by Profit descending
        if profit_col:
            # print("\nDEBUG: Starting sort process")
            profit_idx = ord(profit_col.upper()) - ord('B')
            # print(f"Using profit column index {profit_idx}")
            
            # Print current values before sort
            # print("\nCurrent values in profit column (first 5 rows):")
            for row in range(10, min(15, last_row + 1)):
                cell = ws.cell(row=row, column=ord(profit_col.upper()) - ord('A'))
                # print(f"Row {row}: {cell.value} ({type(cell.value)})")
            
            # Store all row data with formatting
            all_rows = []
            for row in range(10, last_row + 1):
                row_data = []
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    # Store font properties individually rather than the font object
                    font_props = {}
                    if cell.font:
                        font_props = {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'color': cell.font.color,
                            'underline': cell.font.underline,
                            'strike': cell.font.strike,
                            'vertAlign': cell.font.vertAlign
                        }
                    
                    # Store alignment properties individually
                    alignment_props = {}
                    if cell.alignment:
                        alignment_props = {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text,
                            'shrink_to_fit': cell.alignment.shrink_to_fit,
                            'indent': cell.alignment.indent
                        }
                    
                    row_data.append({
                        'value': cell.value if cell.value is not None else 0.0,  # Convert None to 0.0 for sorting
                        'number_format': cell.number_format,
                        'font_props': font_props,
                        'alignment_props': alignment_props,
                        'hyperlink': cell.hyperlink
                    })
                all_rows.append(row_data)
            
            # Sort all rows by profit value
            # print("\nSorting rows...")
            sorted_rows = sorted(
                all_rows,
                key=lambda x: (
                    safe_numeric(x[profit_idx]['value']) or 0.0  # Convert None/NaN to 0.0
                ),
                reverse=True
            )
            
            # print("\nVerifying sort order (first 10 values):")
            # for i, row in enumerate(sorted_rows[:10]):
            #     profit_val = row[profit_idx]['value']
            #     if isinstance(profit_val, float) and math.isnan(profit_val):
            #         profit_val = 0.0
            #     print(f"Position {i}: {profit_val} ({type(profit_val)})")
            
            # print("\nFirst 5 values after sorting:")
            # for i, row in enumerate(sorted_rows[:5]):
            #     print(f"Row {i+10}: {row[profit_idx]['value']} ({type(row[profit_idx]['value'])})")
            
            # Write back sorted data with formatting preserved
            # print("\nWriting back sorted data...")
            from openpyxl.styles import Font, Alignment
            for i, row_data in enumerate(sorted_rows):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=i+10, column=j+2)
                    cell.value = cell_data['value'] if cell_data['value'] != 0.0 else None  # Convert back to None if it was blank
                    cell.number_format = cell_data['number_format']
                    
                    # Create new Font object with stored properties
                    if cell_data['font_props']:
                        cell.font = Font(**cell_data['font_props'])
                    
                    # Create new Alignment object with stored properties
                    if cell_data['alignment_props']:
                        cell.alignment = Alignment(**cell_data['alignment_props'])
                    
                    cell.hyperlink = cell_data['hyperlink']
            
            # 5. Update Profit Rank based on new sort order
            # print("\nUpdating rank...")
            if rank_col:
                for row in range(10, last_row + 1):
                    rank_cell = ws[f'{rank_col}{row}']
                    rank_cell.value = row - 9  # Simple 1-based ranking
                    # if row <= 15:
                    #     print(f"Row {row} new rank: {row - 9}")

    
    header_font = Font(name='Calibri', size=14, bold=True)
    normal_font = Font(name='Calibri', size=12)
    mono_font = Font(name='Consolas', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set initial properties
    ws.sheet_view.zoomScale = 140
    ws['C2'] = address

    # 1. First add Bundle Tx Details columns
    def get_max_bundle_transactions(df, count_column_name):
        try:
            return int(df[count_column_name].max())
        except KeyError:
            return 0

    def add_bundle_tx_details(ws, max_tx, start_col_letter='AX', header_font=None, thin_border=None):
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter, column_index_from_string

        start_col = column_index_from_string(start_col_letter)
        header_row = 9
        header_cell = ws.cell(row=header_row, column=start_col)
        header_cell.value = "Bundle Tx Details"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if header_font:
            header_cell.font = header_font
        if thin_border:
            header_cell.border = thin_border
        
        if max_tx > 1:
            end_col = start_col + max_tx - 1
            ws.merge_cells(
                start_row=header_row, 
                start_column=start_col, 
                end_row=header_row, 
                end_column=end_col
            )

        for i in range(max_tx):
            col = start_col + i
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions[col_letter].alignment = Alignment(horizontal='center')

    def process_transaction_details(ws, df, tx_column_name, searcher_column_name, start_col_letter='AX'):
        from openpyxl.styles import Font, Alignment
        from openpyxl.comments import Comment
        from openpyxl.utils import column_index_from_string

        start_col = column_index_from_string(start_col_letter)

        blue_font = Font(color="0000FF", underline="single")
        green_font = Font(color="008000", underline="single")
        center_align = Alignment(horizontal='center', vertical='center')

        # Set column width for the range
        max_tx_count = 0
        for row in range(10, ws.max_row + 1):
            tx_value = df.iloc[row-10][tx_column_name]
            if pd.notna(tx_value):
                tx_count = len(str(tx_value).split(','))
                max_tx_count = max(max_tx_count, tx_count)

        column_width = 3
        
        # Apply column width to all potentially used columns
        for i in range(max_tx_count):
            col_letter = get_column_letter(start_col + i)
            ws.column_dimensions[col_letter].width = column_width
        
        for row in range(10, ws.max_row + 1):
            tx_value = df.iloc[row-10][tx_column_name]
            searcher_txs = df.iloc[row-10][searcher_column_name]

            if pd.notna(tx_value):
                txs = [tx.strip() for tx in str(tx_value).split(',')]
                searcher_set = set(tx.strip() for tx in str(searcher_txs).split(',')) if pd.notna(searcher_txs) else set()

                for i, tx in enumerate(txs):
                    detail_cell = ws.cell(row=row, column=start_col + i)
                    detail_cell.value = f"Tx{i+1}"
                    detail_cell.hyperlink = f'https://etherscan.io/tx/{tx}'
                    detail_cell.alignment = Alignment(horizontal='center')
                    detail_cell.border = thin_border

                    if tx in searcher_set:
                        detail_cell.font = blue_font
                    else:
                        detail_cell.font = green_font

                    comment = Comment(f"Transaction Hash:\n{tx}", "System")
                    detail_cell.comment = comment

    max_tx = get_max_bundle_transactions(df, "# Txs in Bundle")

    if max_tx > 0:
        add_bundle_tx_details(ws, max_tx, start_col_letter='AX', header_font=header_font, thin_border=thin_border)
        process_transaction_details(ws, df, "Txs in Bundle", "searcher_txs", start_col_letter='AX')

    # 2. Add additional columns
    ws['AT9'] = 'Tenderly'
    ws['AU9'] = 'Eigenphi tx'
    ws['AV9'] = 'Eigenphi block'
    ws['AW9'] = 'libMEV'

    # 3. Move column X to AW (if needed)
    def move_column_to_end(ws, from_col_letter, to_col_letter):
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
        from_col = column_index_from_string(from_col_letter)
        to_col = column_index_from_string(to_col_letter)
        
        # Store column data
        col_data = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=from_col)
            if not isinstance(cell, MergedCell):
                # Get border properties
                border_kwargs = {}
                if cell.border:
                    for side in ['left', 'right', 'top', 'bottom']:
                        side_obj = getattr(cell.border, side)
                        if side_obj:
                            border_kwargs[side] = Side(
                                style=side_obj.style,
                                color=side_obj.color
                            )
                
                # Get fill properties
                fill_kwargs = None
                if cell.fill and cell.fill.fill_type != None:
                    fill_kwargs = {
                        'fill_type': cell.fill.fill_type,
                        'start_color': cell.fill.start_color,
                        'end_color': cell.fill.end_color
                    }
                
                col_data.append({
                    'value': cell.value,
                    'hyperlink': cell.hyperlink,
                    'font_kwargs': {
                        'name': cell.font.name,
                        'size': cell.font.size,
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'color': cell.font.color,
                        'underline': cell.font.underline,
                        'strike': cell.font.strike,
                        'vertAlign': cell.font.vertAlign
                    } if cell.font else None,
                    'alignment_kwargs': {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrap_text': cell.alignment.wrap_text,
                        'shrink_to_fit': cell.alignment.shrink_to_fit,
                        'indent': cell.alignment.indent
                    } if cell.alignment else None,
                    'border_kwargs': border_kwargs,
                    'fill_kwargs': fill_kwargs,
                    'number_format': cell.number_format,
                    'protection': cell.protection.locked if cell.protection else None,
                    'comment': cell.comment
                })
            else:
                col_data.append(None)
        
        # Shift columns left to fill gap
        for row in range(1, ws.max_row + 1):
            for col in range(from_col, to_col):
                source_cell = ws.cell(row=row, column=col + 1)
                target_cell = ws.cell(row=row, column=col)
                if not isinstance(target_cell, MergedCell) and not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value
                    target_cell.hyperlink = source_cell.hyperlink
                    
                    if source_cell.font:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            vertAlign=source_cell.font.vertAlign
                        )
                    
                    if source_cell.alignment:
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text,
                            shrink_to_fit=source_cell.alignment.shrink_to_fit,
                            indent=source_cell.alignment.indent
                        )
                    
                    if source_cell.border:
                        border_kwargs = {}
                        for side in ['left', 'right', 'top', 'bottom']:
                            side_obj = getattr(source_cell.border, side)
                            if side_obj:
                                border_kwargs[side] = Side(
                                    style=side_obj.style,
                                    color=side_obj.color
                                )
                        target_cell.border = Border(**border_kwargs)
                    
                    if source_cell.fill and source_cell.fill.fill_type != None:
                        target_cell.fill = PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                    
                    target_cell.number_format = source_cell.number_format
                    if source_cell.protection:
                        target_cell.protection = Protection(locked=source_cell.protection.locked)
                    target_cell.comment = source_cell.comment
    
        # Write stored data to new position
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=to_col)
            if not isinstance(cell, MergedCell) and col_data[row - 1]:
                data = col_data[row - 1]
                cell.value = data['value']
                cell.hyperlink = data['hyperlink']
                if data['font_kwargs']:
                    cell.font = Font(**data['font_kwargs'])
                if data['alignment_kwargs']:
                    cell.alignment = Alignment(**data['alignment_kwargs'])
                if data['border_kwargs']:
                    cell.border = Border(**data['border_kwargs'])
                if data['fill_kwargs']:
                    cell.fill = PatternFill(**data['fill_kwargs'])
                cell.number_format = data['number_format']
                if data['protection']:
                    cell.protection = Protection(locked=data['protection'])
                cell.comment = data['comment']

    move_column_to_end(ws, 'X', 'AW')

    # 4. Apply all formatting

    # Set background white for all cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = white_fill

    # Format column headers
    for row in ws['B9:' + get_column_letter(ws.max_column) + '9']:
        for cell in row:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Update filter range
    ws.auto_filter.ref = f'B9:{get_column_letter(ws.max_column)}{ws.max_row}'

    # Format data cells
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            if cell.column < 2:
                continue

            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')

            col_letter = get_column_letter(cell.column)
            header_cell = ws[f'{col_letter}9']

            if header_cell.value in ['block #', 'nonce', 'tx Index', 'Profit Rank', 'Resim', 
                                   'Resim @ txIx 0', 'Resim @ txIx -1', 'Tenderly', 
                                   'Eigenphi tx', 'Eigenphi block', 'libMEV',
                                   '# Txs in Bundle', 'Bundle Tx Details', 'searcher_txs_count', 'tokens_count']:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if header_cell.value in ['from', 'to', 'methodId']:
                cell.font = mono_font

            elif header_cell.value in ['tx Hash', 'tx Hash']:
                cell.font = mono_font
                if cell.value:
                    full_hash = str(cell.value)
                    cell.value = str(cell.value)[:10] + '...'
                    cell.hyperlink = f'https://etherscan.io/tx/{full_hash}'
                    cell.font = Font(name='Consolas', size=12, color='0000FF', underline='single')

            elif header_cell.value == 'block #':
                cell.font = mono_font
                if cell.value:
                    block_no = str(cell.value)
                    cell.hyperlink = f'https://etherscan.io/txs?block={block_no}'
                    cell.font = Font(name='Consolas', size=12, color='0000FF', underline='single')

            elif header_cell.value == 'Tenderly':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "tx"
                cell.hyperlink = f'https://dashboard.tenderly.co/tx/mainnet/{full_hash}'

            elif header_cell.value == 'Eigenphi tx':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "tx"
                cell.hyperlink = f'https://eigenphi.io/mev/ethereum/tx/{full_hash}'

            elif header_cell.value == 'Eigenphi block':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "block"
                cell.hyperlink = f'https://eigenphi.io/mev/eigentx/{full_hash}?tab=block'

            elif header_cell.value == 'libMEV':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "block"
                cell.hyperlink = f'https://libmev.com/blocks/{block_no}'

            elif header_cell.value in ['Resim @ txIx 0', 'Resim @ txIx -1']:
                if cell.value == 'OK':
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

            elif header_cell.value == 'dateTime':
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Format summary section
    def format_summary_headers(cell):
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(size=14, bold=True)

    def format_summary_values(cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=14, bold=False)

    # Add and format summary headers
    summary_headers = {
        'C4': 'Start Block:', 'C5': 'End Block:', 'C6': 'Total Transactions:', 'C7': 'Total Profit:',
        'G4': 'Start Balance:', 'G5': 'End Balance:', 'G6': 'Withdrawals:', 'G7': 'Total Margin:'
    }

    for pos, text in summary_headers.items():
        ws[pos] = text
        format_summary_headers(ws[pos])

    # Add formulas
    formulas = {
        'D4': f'=MIN(B10:B{len(df)+7})',
        'D5': f'=MAX(B10:B{len(df)+7})',
        'D6': f'=COUNTA(B10:B{len(df)+7})',
        'D7': f'=SUM(H5-H4+H6)',
        'H4': f'=VLOOKUP(D4,B10:H{len(df)+7},7,FALSE)',
        'H5': f'=VLOOKUP(D5,B10:H{len(df)+7},7,FALSE)',
        'H6': f'=SUMIF(I10:I{len(df)+9},"Withdrawal",Q10:Q{len(df)+9})',
        'H7': f'=TEXT(ROUND((1-(SUM(N10:O{len(df)+9})/SUM(P10:P{len(df)+9}))), 5), "0.0%")'
    }

    for pos, formula in formulas.items():
        ws[pos] = formula
        format_summary_values(ws[pos])

    if address == "0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D":
            special_formulas = {
                'D7': f'=ROUND(SUM($P$10:$P${len(df)+9})+$H$6,3)',  # Added $ for absolute references
            }
            for pos, formula in special_formulas.items():
                ws[pos] = formula
                format_summary_values(ws[pos])

    # After the header formatting loop where you set font, alignment, and border
    for row in ws['B9:' + get_column_letter(ws.max_column) + '9']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set row height to auto
    ws.row_dimensions[9].height = None  # This triggers auto-height calculation

    # Format address header
    ws.merge_cells('C2:G2')
    cell = ws['C2']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=16, bold=True)

    # 5. Apply column widths
    width_groups = {
        20: ['A', 'D'],
        17.5: ['C', 'G'],
        15: ['B'],
        10: ['H', 'O', 'P', 'U', 'V', 'X', 'AS', 'AT'],
        8.5: ['E', 'F', 'S', 'AW'],
        7.5: ['Q', 'S', 'T', 'AU', 'AV'],
        5: ['AX', 'AY', 'AZ']
    }

    for width, columns in width_groups.items():
        for col in columns:
            ws.column_dimensions[col].width = width

    # 6. Apply grouping and hiding (LAST)
    def get_columns_in_range(start_col, end_col):
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

    # Group and hide columns
    for col_letter in get_columns_in_range('I', 'P'):
        ws.column_dimensions[col_letter].outline_level = 1
        ws.column_dimensions[col_letter].hidden = True

    for col_letter in get_columns_in_range('W', 'AR'):
        ws.column_dimensions[col_letter].outline_level = 1
        ws.column_dimensions[col_letter].hidden = True

    # Final formatting tweaks
    # Alternate colors only from column B
    for row in range(10, ws.max_row + 1):
        if (row - 9) % 2 == 0:
            for cell in ws[row]:
                if 2 <= cell.column <= len(df.columns) + 2:  # Only B onwards
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # Add profit column color scale
    profit_col = get_column_letter(df.columns.get_loc('Profit') + 2)  # +2 for startcol offset
    ws.conditional_formatting.add(
        f'{profit_col}10:{profit_col}{ws.max_row}',
        ColorScaleRule(
            start_type='min',
            start_color='FF0000',
            mid_type='percentile', 
            mid_value=50,
            mid_color='FFFF00',
            end_type='max',
            end_color='00FF00'
        )
    )

    ws.merge_cells('AT9:AU9')
    ws['AT9'] = "Eigenphi"

    def format_bundle_tx_details(ws):
        """
        Format all columns under 'Bundle Tx Details' with consistent styling
        Groups columns after the first 5
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font
    
        # Find the Bundle Tx Details header (should be in row 9)
        start_col = None
        end_col = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=9, column=col)
            if cell.value == "Bundle Tx Details":
                if start_col is None:
                    start_col = col
                end_col = col
    
        if start_col is None:
            print("Bundle Tx Details header not found")
            return
    
        # Get last row with data
        last_row = ws.max_row
    
        # Create style objects
        alignment = Alignment(horizontal='center', vertical='center')
        font = Font(size=8, color="0000FF", underline="single")
        column_width = 2.5
    
        # Count how many columns are in the Bundle Tx Details section
        bundle_cols = []
        for col in range(start_col, ws.max_column + 1):
            # Check if we've moved past the Bundle Tx Details section
            header_value = ws.cell(row=9, column=col).value
            if header_value and header_value != "Bundle Tx Details":
                break
            bundle_cols.append(col)
    
        total_bundle_cols = len(bundle_cols)
        # print(f"DEBUG: Found {total_bundle_cols} columns in Bundle Tx Details")
    
        # Format all columns
        for col in bundle_cols:
            col_letter = get_column_letter(col)
            
            # Set column width
            ws.column_dimensions[col_letter].width = column_width
            
            # Apply formatting to cells
            for row in range(10, last_row + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = alignment
                    cell.font = font
                except:
                    pass
        
            # Group columns after the first 5
            if col >= start_col + 5:  # If this is the 6th or later column
                ws.column_dimensions[col_letter].outline_level = 1
                ws.column_dimensions[col_letter].hidden = True  # Initially hidden
    
        # If we have more than 5 columns, make sure outlining is enabled
        if total_bundle_cols > 5:
            ws.sheet_view.showOutlineSymbols = True

    format_bundle_tx_details(ws)

    def add_bribe_percentage(ws):
        """
        Insert a new 'Bribe %' column after 'Margin %' and group them
        Uses Excel formulas to calculate Bribe % as (1 - Margin %)
        Completely matches formatting from Margin % (percentage format, borders, shading)
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        
        # Find the 'Margin %' column
        margin_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=9, column=col).value == "Margin %":
                margin_col = col
                break
                
        if margin_col is None:
            print("Margin % column not found")
            return
            
        # Get the last row with data
        last_row = ws.max_row
        
        # Insert a new column after Margin %
        ws.insert_cols(idx=margin_col + 1)
        
        # Get formatting from margin header cell
        margin_header = ws.cell(row=9, column=margin_col)
        
        # Set the header for the new column (create new style objects)
        bribe_header = ws.cell(row=9, column=margin_col + 1)
        bribe_header.value = "Bribe %"
        
        # Create new Font object
        if margin_header.font:
            bribe_header.font = Font(
                name=margin_header.font.name,
                size=margin_header.font.size,
                bold=margin_header.font.bold,
                italic=margin_header.font.italic,
                color=margin_header.font.color
            )
        
        # Create new Alignment object
        bribe_header.alignment = Alignment(horizontal='center', vertical='center')
        
        # Create new Border object if needed
        if margin_header.border:
            border_sides = {}
            for side in ['left', 'right', 'top', 'bottom']:
                side_obj = getattr(margin_header.border, side)
                if side_obj and side_obj.style:
                    border_sides[side] = Side(style=side_obj.style, color=side_obj.color)
            
            if border_sides:
                bribe_header.border = Border(**border_sides)
        
        # Copy Fill (background shading)
        if margin_header.fill and margin_header.fill.fill_type != 'none':
            bribe_header.fill = PatternFill(
                fill_type=margin_header.fill.fill_type,
                start_color=margin_header.fill.start_color,
                end_color=margin_header.fill.end_color
            )
        
        # Get margin column letter for formulas
        margin_col_letter = get_column_letter(margin_col)
        
        # Add formula for each row: Bribe % = 1 - Margin %
        for row in range(10, last_row + 1):
            margin_cell = ws.cell(row=row, column=margin_col)
            bribe_cell = ws.cell(row=row, column=margin_col + 1)
            
            # Set formula: =1-[Margin %]
            bribe_cell.value = f"=1-{margin_col_letter}{row}"
            
            # Ensure percentage formatting
            if "%" in margin_cell.number_format:
                bribe_cell.number_format = margin_cell.number_format
            else:
                # If margin is not formatted as percentage, apply a percentage format
                bribe_cell.number_format = "0.0%"
            
            # Create new Font object if needed
            if margin_cell.font:
                bribe_cell.font = Font(
                    name=margin_cell.font.name,
                    size=margin_cell.font.size,
                    bold=margin_cell.font.bold,
                    italic=margin_cell.font.italic,
                    color=margin_cell.font.color
                )
            
            # Create new Alignment object
            if margin_cell.alignment:
                bribe_cell.alignment = Alignment(
                    horizontal=margin_cell.alignment.horizontal,
                    vertical=margin_cell.alignment.vertical,
                    wrap_text=margin_cell.alignment.wrap_text
                )
                
            # Create Border object
            if margin_cell.border:
                border_sides = {}
                for side in ['left', 'right', 'top', 'bottom']:
                    side_obj = getattr(margin_cell.border, side)
                    if side_obj and side_obj.style:
                        border_sides[side] = Side(style=side_obj.style, color=side_obj.color)
                
                if border_sides:
                    bribe_cell.border = Border(**border_sides)
            
            # Copy Fill (background shading)
            if margin_cell.fill and margin_cell.fill.fill_type != 'none':
                bribe_cell.fill = PatternFill(
                    fill_type=margin_cell.fill.fill_type,
                    start_color=margin_cell.fill.start_color,
                    end_color=margin_cell.fill.end_color
                )
        
        # Group the columns and hide Margin %
        ws.column_dimensions.group(margin_col_letter, margin_col_letter, hidden=True)
        
        # Set column width of Bribe % to match Margin % width
        bribe_col_letter = get_column_letter(margin_col + 1)
        ws.column_dimensions[bribe_col_letter].width = ws.column_dimensions[margin_col_letter].width

    add_bribe_percentage(ws)

    # Set row height
    ws.row_dimensions[10].height = 20

    # Set worksheet title
    ws.title = address[:8]

    # Save workbook
    wb.save(output_xlsx)


if __name__ == "__main__":

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Process all addresses once, storing the first latest_block for filename
    first_run = True
    output_xlsx = None

    for addr in address_list:

        # if addr == "0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D":
        if addr in ["0x1b9FcB24c533839dC847235bd8Eb80E37EC42f85",]:
            start_block_adj = start_block
        elif addr in ["0x0BdE59981FDEaC219Ce9E618d27F193438Bff786"]:
            start_block_adj = start_block + 3000
        elif addr in ["0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D"]:
            start_block_adj = start_block + 3000
        else:
            start_block_adj = start_block

        combined_df, merged_df, libmev_df, final_df, latest_block = process_address(addr, start_block_adj)

        if first_run:
            output_xlsx = f'MEV_daily_multiple_addresses_{latest_block}_P&L.xlsx'
            first_run = False

        # Create new worksheet for this address
        ws = wb.create_sheet(title=addr[:8])
        # format_pl_report(addr, merged_df, output_xlsx, wb, ws)
        format_pl_report(addr, final_df, output_xlsx, wb, ws)

    # Save the workbook once at the end
    wb.save(output_xlsx)

    print("="*20)
    print("ALL DONE NOW!")
    print("="*20, "\n")
