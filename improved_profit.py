import os
import time
import json
import math
import requests
import backoff
import logging
import asyncio
import aiohttp
import nest_asyncio
import subprocess
import tempfile
import pandas as pd
import numpy as np
import argparse
import importlib.util
import sys

from datetime import datetime, timezone, timedelta
from pathlib import Path
from web3 import Web3
from decimal import Decimal
from eth_typing import HexStr
from hexbytes import HexBytes
from tqdm import tqdm  # Using regular tqdm instead of tqdm.notebook
from PIL import Image
from dotenv import load_dotenv

# Project imports
from get_profit_of_wallet import (
    trace_list_generator,
    tx_list_generator,
    get_address_tx_hashes_and_blocks,
    receipt_list_generator,
    block_list_generator,
    process_initial_dataframe,
    libmev_data,
    combine_blockchain_data,
    merge_dataframes,
    create_final_df
)
from lib_etherscan_funcs import get_address_tx_hashes_and_blocks
from chain_lib import w3_deejmon_http, deejmon_http, chain_data, minimal_abi
from metadata import miner_map, token_contracts, weth_contract, usdt_contract, usdc_contract, stETH_contract, entity_groups
from block_calculation_engine import get_timestamps, fetch_block_number, main

# Excel formatting imports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("mev_analysis.log")
    ]
)
logger = logging.getLogger(__name__)

# Try to import the wallet grouping utility
# try:
#     from wallet_grouping import detect_related_wallets
#     wallet_grouping_available = True
# except ImportError:
#     wallet_grouping_available = False


# Load environment variables
load_dotenv()

# Connect to Web3
W3 = w3_deejmon_http

# Create reverse mapping for quick lookup
address_to_entity = {}
for entity, addresses in entity_groups.items():
    for addr in addresses:
        address_to_entity[addr.lower()] = entity

# Constants
MAX_BATCH_SIZE = 200  # Maximum items per batch
RETRY_COUNT = 3  # Number of retries for failed operations

# Helper functions
def connect_with_retries(web3, retries=5, delay=0.1):
    """Connect to Web3 provider with retries"""
    for attempt in range(retries):
        if web3.is_connected():
            logger.info("Connected to Ethereum node!")
            return True
        else:
            logger.warning(f"Connection attempt {attempt + 1} failed. Retrying in {delay} seconds...")
            time.sleep(delay)
    return False


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='MEV Profit Analysis Tool')
    parser.add_argument('--time_period', type=str, default='yesterday',
                      choices=['current_datetime', 'yesterday', 'midnight', 
                               'yesterday_midnight', 'start_of_week', 'start_of_month', 'custom'],
                      help='Time period to analyze')
    parser.add_argument('--addresses', type=str, 
                      help='Comma-separated list of addresses (overrides .env)')
    parser.add_argument('--batch_size', type=int, default=MAX_BATCH_SIZE,
                      help='Maximum batch size for API requests')
    parser.add_argument('--output', type=str, default=None,
                      help='Output Excel file path')

    # Add custom date range options
    parser.add_argument('--start_date', type=str, 
                      help='Start date for custom range (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, default=datetime.now().strftime('%Y-%m-%d'),
                      help='End date for custom range (YYYY-MM-DD), defaults to today)')

    return parser.parse_args()


def get_abi_from_etherscan(contract_address, etherscan_api_key):
    """Fetch contract ABI from Etherscan"""
    url = f"https://api.etherscan.io/api?module=contract&action=getabi&address={contract_address}&apikey={etherscan_api_key}"
    try:
        response = requests.get(url)
        data = response.json()
        if data["status"] == "1":
            return data["result"]
        logger.warning(f"Failed to get ABI: {data.get('message')}")
        return None
    except Exception as e:
        logger.error(f"Error fetching ABI: {str(e)}")
        return None


def chunk_list(lst, chunk_size):
    """Split a list into chunks of specified size"""
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]


async def process_in_batches(items, process_func, batch_size=MAX_BATCH_SIZE, desc="Processing"):
    """Process items in batches with progress bar"""
    results = []
    batches = chunk_list(items, batch_size)
    
    with tqdm(total=len(items), desc=desc, ncols=100) as pbar:
        for batch in batches:
            batch_results = await process_func(batch)
            results.extend(batch_results)
            pbar.update(len(batch))
    
    return results


def set_sheet_white(ws):
    """Set entire sheet to white background with no grid lines"""
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Apply to all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = white_fill
    
    # Turn off grid lines
    ws.sheet_view.showGridLines = False


# Main data processing functions
def get_token_balances_with_retry(W3, address, block_list, token_contracts, max_retries=3, batch_size=MAX_BATCH_SIZE):
    """Get token balances with retry logic and improved batching"""
    retry_count = 0
    all_results = []
    
    # Split block list into manageable batches
    block_batches = chunk_list(block_list, batch_size)
    
    while retry_count < max_retries:
        try:
            eth_results = []
            weth_results = []
            erc20_results_dict = {token: [] for token in token_contracts if token != 'WETH'}
            
            # Process each batch
            for batch_idx, batch_blocks in enumerate(block_batches):
                batch_desc = f"Batch {batch_idx+1}/{len(block_batches)}"
                
                with tqdm(total=len(batch_blocks), desc=f"Getting balances: {batch_desc}", ncols=100, leave=False) as pbar:
                    # Get ETH balances
                    eth_batch = W3.batch_requests()
                    for block in batch_blocks:
                        eth_batch.add(W3.eth.get_balance(address, block_identifier=int(block)))
                    eth_batch_results = eth_batch.execute()
                    eth_results.extend(eth_batch_results)
                    pbar.update(len(batch_blocks) // 3)
                    
                    # Get WETH balances
                    weth_batch = W3.batch_requests()
                    for block in batch_blocks:
                        weth_batch.add(token_contracts['WETH'].functions.balanceOf(address).call(block_identifier=int(block)))
                    weth_batch_results = weth_batch.execute()
                    weth_results.extend(weth_batch_results)
                    pbar.update(len(batch_blocks) // 3)
                    
                    # Get other token balances
                    for token, contract in token_contracts.items():
                        if token == 'WETH':
                            continue
                        
                        erc20_batch = W3.batch_requests()
                        for block in batch_blocks:
                            erc20_batch.add(contract.functions.balanceOf(address).call(block_identifier=int(block)))
                        
                        token_balances = erc20_batch.execute()
                        erc20_results_dict[token].extend(token_balances)
                    
                    pbar.update(len(batch_blocks) // 3)
                
                logger.info(f"✅ Processed balances for batch {batch_idx+1}/{len(block_batches)}")
            
            # Combine results
            return combine_token_balances(address, block_list, eth_results, weth_results, erc20_results_dict)
            
        except Exception as e:
            retry_count += 1
            logger.error(f"Balance retrieval attempt {retry_count} failed: {str(e)}")
            if retry_count < max_retries:
                logger.info("Waiting before retry...")
                time.sleep(5)
                # Reduce batch size on failure
                batch_size = max(batch_size // 2, 10)
                block_batches = chunk_list(block_list, batch_size)
                logger.info(f"Reduced batch size to {batch_size}")
            else:
                logger.error("Max retries reached for balance retrieval")
                return []

def combine_token_balances(address, block_list, eth_results, weth_results, erc20_results_dict):
    """
    Combine ETH, WETH, and other ERC20 token balances into a structured format
    """
    # Ensure all lists are the same length
    if not (len(block_list) == len(eth_results) == len(weth_results)):
        logger.error(f"Mismatched lengths: blocks={len(block_list)}, eth={len(eth_results)}, weth={len(weth_results)}")
        raise ValueError("Mismatched lengths between blocks and results")

    # Also verify that each token in erc20_results_dict has the same length
    for token, balances in erc20_results_dict.items():
        if len(balances) != len(block_list):
            logger.error(f"Mismatched lengths for token {token}: {len(balances)} vs {len(block_list)}")
            raise ValueError(f"Mismatched lengths for token {token}")

    combined_balances = []

    # Create a list of token names for consistent ordering
    token_names = sorted(erc20_results_dict.keys())

    # Zip all lists together and create dictionaries
    for i, (block, eth_bal, weth_bal) in enumerate(zip(block_list, eth_results, weth_results)):
        balance_dict = {
            "wallet_address": address,
            "block_number": int(block),
            "ETH": int(eth_bal),
            "WETH": int(weth_bal)
        }

        # Add other ERC20 token balances
        for token in token_names:
            balance_dict[token] = int(erc20_results_dict[token][i])

        combined_balances.append(balance_dict)

    return combined_balances

async def process_transactions(tx_results_dicts, deejmon_http, batch_size=MAX_BATCH_SIZE):
    """Process transactions with improved async batching"""
    async with aiohttp.ClientSession() as session:
        # Split into batches
        batches = chunk_list(tx_results_dicts, batch_size)
        all_results = []
        
        for batch_idx, batch in enumerate(batches):
            batch_tasks = []
            for count, tx in enumerate(batch):
                resim_bundle = {
                    "from": tx['from'],
                    "to": tx['to'],
                    "data": tx['input'].hex() if hasattr(tx['input'], 'hex') else tx['input'],
                    "gas": tx['gas'],
                    "value": tx['value'],
                    "blockNumber": tx['blockNumber'],
                }

                # Create tasks in order but don't await them yet
                task0 = resim_tx_async(session, deejmon_http, resim_bundle, tx['transactionIndex'])
                task1 = resim_tx_async(session, deejmon_http, resim_bundle, 0)
                task2 = resim_tx_async(session, deejmon_http, resim_bundle, -1)
                batch_tasks.append((count, [task0, task1, task2]))  # Keep track of original order

            # Process this batch
            batch_results = []
            with tqdm(total=len(batch_tasks), desc=f"Processing transactions batch {batch_idx+1}/{len(batches)}", ncols=100) as pbar:
                for count, task_group in batch_tasks:
                    group_results = await asyncio.gather(*task_group)
                    batch_results.extend(group_results)
                    pbar.update(1)
            
            all_results.extend(batch_results)
        
        # Reorder and process results
        processed_results = []
        for i in range(0, len(all_results), 3):
            tx_index = i // 3
            
            # Handle index errors gracefully
            if tx_index < len(tx_results_dicts):
                if i+2 < len(all_results):
                    result_orig = all_results[i]
                    result_0 = all_results[i + 1]
                    result_n1 = all_results[i + 2]
                    
                    tx_results_dicts[tx_index]['sim_result_orig'] = parse_sim_result(result_orig['result'])
                    tx_results_dicts[tx_index]['sim_result_0'] = parse_sim_result(result_0['result'])
                    tx_results_dicts[tx_index]['sim_result_n1'] = parse_sim_result(result_n1['result'])
                    
                    processed_results.append(tx_results_dicts[tx_index])
        
        return processed_results

async def resim_tx_async(session, deejmon_http, resim_bundle, tx_index):
    """Submit async request to resimulate a transaction"""
    payload = {
        "id": 1,
        "jsonrpc": "2.0",
        "method": "debug_traceCallMany",
        "params": [
            [
                {
                    "blockOverride": {},
                    "transactions": [
                        {
                            "from": resim_bundle['from'],
                            "to": resim_bundle['to'],
                            "data": resim_bundle['data'],
                            "gas": resim_bundle['gas'],
                            "value": resim_bundle['value'],
                        }
                    ]
                }
            ],
            {
                "blockNumber": hex(resim_bundle['blockNumber']),
                "transactionIndex": tx_index,
            },
            {
                "tracer": "callTracer",
                "tracerConfig": {
                    "diffMode": False,
                    "onlyTopCall": True,
                    "withLog": False
                }
            }
        ]
    }
    headers = {
        "accept": "application/json",
        "content-type": "application/json"
    }

    # Use exponential backoff for retries
    @backoff.on_exception(
        backoff.expo,
        (aiohttp.ClientError, asyncio.TimeoutError, ConnectionError),
        max_tries=3
    )
    async def request_with_retry():
        async with session.post(deejmon_http, json=payload, headers=headers, timeout=30) as response:
            return await response.json()

    try:
        result = await request_with_retry()
        return {
            'tx_index': tx_index,
            'resim_bundle': resim_bundle,
            'result': result
        }
    except Exception as e:
        logger.error(f"Error in resim_tx_async: {str(e)}")
        return {
            'tx_index': tx_index,
            'resim_bundle': resim_bundle,
            'result': {"error": {"message": str(e)}}
        }

def parse_sim_result(result):
    """Parse simulation result with improved error handling"""
    if result is None:
        return "Error:NoResult"
        
    if "error" in result:
        error_msg = result.get("error", {}).get("message", "UnknownError")
        return f"Error:{error_msg}"
        
    try:
        if "result" in result:
            if isinstance(result["result"], list) and len(result["result"]) > 0:
                if len(result["result"][0]) > 0:
                    if "error" in result["result"][0][0]:
                        if "revertReason" in result["result"][0][0]:
                            return f"Revert:{result['result'][0][0]['revertReason']}"
                        else:
                            return f"Error:{result['result'][0][0]['error']}"
        return "OK"
    except (KeyError, IndexError, TypeError) as e:
        logger.error(f"Error parsing simulation result: {str(e)}")
        return f"Error:ParsingFailed"

def process_address(address, start_block, batch_size=MAX_BATCH_SIZE):
    """Process a single address with improved batching and error handling"""
    try:
        address = Web3.to_checksum_address(address.lower())
        logger.info(f"Processing address: {address}")
        
        # Get transaction history
        L24_tx = get_address_tx_hashes_and_blocks(address, start_block)
        if not L24_tx:
            logger.warning(f"No transactions found for address {address} from block {start_block}")
            return None, None, None, None, start_block
            
        df = pd.DataFrame(L24_tx)
        
        # Get ABI
        etherscan_api_key = os.environ.get("etherscan_api_key")
        usdc_proxy = Web3.to_checksum_address("0x43506849d7c04f9138d1a2050bbf3a0c054402dd")
        abi = get_abi_from_etherscan(usdc_proxy, etherscan_api_key)
        
        # Prepare data
        tx_list = list(set([x['txHash'] for x in L24_tx]))
        unique_blocks = list(set(df['blockNumber'].to_list()))
        latest_block = max(unique_blocks)
        
        # Get token balances with improved batching
        balances = get_token_balances_with_retry(
            W3, address, unique_blocks, token_contracts, 
            max_retries=RETRY_COUNT, batch_size=batch_size
        )
        
        # Process blocks in batches
        logger.info(f"Fetching {len(unique_blocks)} unique blocks")
        block_batches = chunk_list(unique_blocks, batch_size)
        block_results = []
        
        with tqdm(total=len(unique_blocks), desc="Fetching blocks", ncols=100) as pbar:
            for batch in block_batches:
                batch_request = W3.batch_requests()
                for block in batch:
                    batch_request.add(W3.eth.get_block(int(block)))
                results = batch_request.execute()
                block_results.extend(results)
                pbar.update(len(batch))

        # Process block data
        block_miners = {block.number: block.miner for block in block_results}
        block_details = block_list_generator(block_results)
        logger.info(f"✅ Fetched blocks")
        
        # Process transactions in batches
        logger.info(f"Fetching {len(tx_list)} transactions")
        tx_batches = chunk_list(tx_list, batch_size)
        tx_results = []
        
        with tqdm(total=len(tx_list), desc="Fetching transactions", ncols=100) as pbar:
            for batch in tx_batches:
                batch_request = W3.batch_requests()
                for tx in batch:
                    batch_request.add(W3.eth.get_transaction(tx))
                batch_results = batch_request.execute()
                tx_results.extend(batch_results)
                pbar.update(len(batch))
        
        logger.info(f"✅ Fetched transactions")
        
        # Process receipts in batches
        logger.info(f"Fetching {len(tx_list)} transaction receipts")
        receipt_batches = chunk_list(tx_list, batch_size)
        receipt_results = []
        
        with tqdm(total=len(tx_list), desc="Fetching transaction receipts", ncols=100) as pbar:
            for batch in receipt_batches:
                batch_request = W3.batch_requests()
                for tx in batch:
                    batch_request.add(W3.eth.get_transaction_receipt(tx))
                batch_results = batch_request.execute()
                receipt_results.extend(batch_results)
                pbar.update(len(batch))
        
        receipt_details = receipt_list_generator(receipt_results)
        logger.info(f"✅ Fetched transaction receipts")
        
        # Process trace data in batches
        logger.info(f"Fetching {len(tx_list)} transaction traces")
        trace_batches = chunk_list(tx_list, batch_size)
        trace_results = []
        
        with tqdm(total=len(tx_list), desc="Fetching transaction traces", ncols=100) as pbar:
            for batch in trace_batches:
                batch_request = W3.batch_requests()
                for tx_hash in batch:
                    tx_hash = tx_hash.hex() if isinstance(tx_hash, HexBytes) else tx_hash
                    batch_request.add(W3.tracing.trace_transaction(tx_hash))
                batch_results = batch_request.execute()
                trace_results.extend(batch_results)
                pbar.update(len(batch))
        
        trace_details = trace_list_generator(block_results, trace_results, block_miners)
        logger.info(f"✅ Fetched transaction traces")
        
        # Use asyncio to process tx resimulations
        logger.info("Processing transaction simulations")
        nest_asyncio.apply()
        tx_results_dicts = [dict(tx) for tx in tx_results]
        updated_tx_results = asyncio.run(process_transactions(tx_results_dicts, deejmon_http, batch_size))
        tx_details = tx_list_generator(updated_tx_results)
        
        # Process dataframes
        logger.info("Processing data into dataframes")
        initial_df = process_initial_dataframe(df, address, balances)
        combined_data = combine_blockchain_data(block_details, tx_details, receipt_details, trace_details, balances)
        combined_df = pd.DataFrame(combined_data)
        
        # Get libMEV data
        logger.info("Fetching libMEV data")
        libmev_df = libmev_data(combined_df)
        
        # Merge dataframes
        merged_df = merge_dataframes(initial_df, combined_df, libmev_df)
        final_df = create_final_df(libmev_df, merged_df)
        
        logger.info(f"Completed processing for address {address}")
        return combined_df, merged_df, libmev_df, final_df, latest_block
        
    except Exception as e:
        logger.error(f"Error processing address {address}: {str(e)}", exc_info=True)
        return None, None, None, None, start_block

def create_summary_leaderboard(all_results, wb):
    """Create a summary leaderboard sheet from all processed addresses"""
    # Create a new sheet for the leaderboard
    ws = wb.create_sheet("Leaderboard")
    set_sheet_white(ws)
    
    # Prepare data for the leaderboard
    summary_data = []
    entity_data = {}
    
    for address, result in all_results.items():
        if not all(r is not None for r in result[:4]):
            continue
            
        combined_df, merged_df, libmev_df, final_df, latest_block = result
        
        # Extract key metrics
        total_txs = len(final_df) if final_df is not None else 0
        total_profit = final_df['Profit'].sum() if final_df is not None else 0
        avg_margin = final_df['Margin %'].str.rstrip('%').astype(float).mean() if final_df is not None else 0
        max_profit_tx = final_df['Profit'].max() if final_df is not None else 0
        
        # Find entity if it exists
        entity = address_to_entity.get(address.lower(), "Individual")
        
        # Add to summary data
        summary_data.append({
            'Address': address,
            'Entity': entity,
            'Total Transactions': total_txs,
            'Total Profit (ETH)': total_profit,
            'Average Margin (%)': avg_margin,
            'Largest Profit (ETH)': max_profit_tx
        })
        
        # Aggregate by entity
        if entity not in entity_data:
            entity_data[entity] = {
                'Total Transactions': 0,
                'Total Profit (ETH)': 0,
                'Addresses': [],
                'Max Profit TX': 0
            }
        
        entity_data[entity]['Total Transactions'] += total_txs
        entity_data[entity]['Total Profit (ETH)'] += total_profit
        entity_data[entity]['Addresses'].append(address)
        entity_data[entity]['Max Profit TX'] = max(entity_data[entity]['Max Profit TX'], max_profit_tx)
    
    # Create a DataFrame for the summary
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        
        # Sort by total profit
        summary_df = summary_df.sort_values('Total Profit (ETH)', ascending=False)
        
        # Format for Excel
        summary_df['Total Profit (ETH)'] = summary_df['Total Profit (ETH)'].round(4)
        summary_df['Average Margin (%)'] = summary_df['Average Margin (%)'].round(2)
        summary_df['Largest Profit (ETH)'] = summary_df['Largest Profit (ETH)'].round(4)
        
        # Add to spreadsheet
        # Add title
        ws.merge_cells('B2:G2')
        ws['B2'] = 'MEV Profit Analysis Leaderboard'
        ws['B2'].font = Font(size=20, bold=True)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add timestamp
        ws['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['B3'].font = Font(size=12, italic=True)
        
        # Add individual wallet section
        ws['B5'] = 'Wallet Performance'
        ws['B5'].font = Font(size=16, bold=True)
        
        # Add DataFrame data starting at row 7
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False), 7):
            for c_idx, value in enumerate(row, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format cells
                if c_idx == 2:  # Address column
                    cell.font = Font(name='Consolas', size=10)
                    cell.alignment = Alignment(horizontal='left')
                    
                    # Add hyperlink to sheet
                    if isinstance(value, str) and len(value) > 2:
                        try:
                            cell.hyperlink = f"#{value[:8]}!A1"
                            cell.font = Font(name='Consolas', size=10, color='0000FF', underline='single')
                        except:
                            pass
                            
                elif c_idx > 3:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # Add headers with bold font
        for col in range(2, len(summary_df.columns) + 2):
            ws.cell(row=6, column=col).font = Font(bold=True)
            ws.cell(row=6, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=6, column=col).border = Border(bottom=Side(style='medium'))
        
        # Column widths
        ws.column_dimensions['B'].width = 40  # Address
        ws.column_dimensions['C'].width = 20  # Entity
        for col in ['D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        
        # Create charts
        create_leaderboard_charts(ws, summary_df, entity_data)
    
    else:
        # Add message if no data
        ws['B2'] = 'No data available for leaderboard'
        ws['B2'].font = Font(size=14, bold=True)
    
    return ws

def create_leaderboard_charts(ws, summary_df, entity_data):
    """Create charts for the leaderboard summary page"""
    # Entity Profit Comparison Chart (Pie Chart)
    entity_profit = pd.DataFrame([
        {'Entity': entity, 'Total Profit (ETH)': data['Total Profit (ETH)']}
        for entity, data in entity_data.items()
    ])
    
    if not entity_profit.empty:
        pie = PieChart()
        pie.title = "Profit Distribution by Entity"
        
        # Add data
        data_rows = len(entity_profit) + 1  # +1 for header
        values = Reference(ws, min_col=5, min_row=6, max_row=6+data_rows)  # Total Profit column
        categories = Reference(ws, min_col=3, min_row=7, max_row=6+data_rows)  # Entity column
        
        pie.add_data(values, titles_from_data=True)
        pie.set_categories(categories)
        
        # Format and position
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.height = 8
        pie.width = 10
        
        # Add to worksheet
        ws.add_chart(pie, "B15")
    
    # Top Wallets by Profit Chart (Bar Chart)
    if len(summary_df) > 0:
        chart = BarChart()
        chart.title = "Top Wallets by Profit"
        chart.type = "col"
        chart.style = 10
        
        # Use only top 5 wallets for the chart
        top_wallets = min(5, len(summary_df))
        
        # Add data
        values = Reference(ws, min_col=5, min_row=6, max_row=6+top_wallets)  # Total Profit column
        categories = Reference(ws, min_col=2, min_row=7, max_row=6+top_wallets)  # Address column
        
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        
        # Format and position
        chart.y_axis.title = "Profit (ETH)"
        chart.x_axis.title = "Address"
        chart.height = 8
        chart.width = 15
        
        # Add to worksheet
        ws.add_chart(chart, "J15")
    
    # Entity summary data on sheet
    # Add entity summary section
    row_start = 7 + len(summary_df) + 3  # Leave gap after wallet data
    
    ws.cell(row=row_start, column=2, value="Entity Summary").font = Font(size=16, bold=True)
    row_start += 2
    
    # Headers
    headers = ["Entity", "Wallets", "Total Transactions", "Total Profit (ETH)", "Largest Profit (ETH)"]
    for i, header in enumerate(headers):
        ws.cell(row=row_start, column=i+2, value=header).font = Font(bold=True)
        ws.cell(row=row_start, column=i+2).alignment = Alignment(horizontal='center')
        ws.cell(row=row_start, column=i+2).border = Border(bottom=Side(style='medium'))
    
    # Entity data
    row = row_start + 1
    for entity, data in sorted(entity_data.items(), key=lambda x: x[1]['Total Profit (ETH)'], reverse=True):
        ws.cell(row=row, column=2, value=entity)
        ws.cell(row=row, column=3, value=len(data['Addresses']))
        ws.cell(row=row, column=4, value=data['Total Transactions'])
        ws.cell(row=row, column=5, value=round(data['Total Profit (ETH)'], 4))
        ws.cell(row=row, column=6, value=round(data['Max Profit TX'], 4))
        
        # Format
        ws.cell(row=row, column=2).font = Font(bold=True)
        for col in range(3, 7):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        row += 1

def create_dashboard_sheet(wb, all_results):
    """Create an executive dashboard summary sheet"""
    ws = wb.create_sheet("Dashboard", 0)  # Add at the beginning
    set_sheet_white(ws)
    
    # Title
    ws.merge_cells('B2:L2')
    title_cell = ws['B2']
    title_cell.value = "MEV PROFIT ANALYSIS DASHBOARD"
    title_cell.font = Font(name='Arial', size=24, bold=True, color='4472C4')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle with date range
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    ws.merge_cells('B3:L3')
    ws['B3'] = f"Analysis Period: {yesterday} - {today}"
    ws['B3'].font = Font(size=12, italic=True)
    ws['B3'].alignment = Alignment(horizontal='center')
    
    # KPI Section
    ws.merge_cells('B5:L5')
    ws['B5'] = "KEY PERFORMANCE INDICATORS"
    ws['B5'].font = Font(size=14, bold=True)
    ws['B5'].alignment = Alignment(horizontal='center')
    
    # Create KPI boxes
    kpi_metrics = calculate_kpi_metrics(all_results)
    
    create_kpi_box(ws, 'B7', 'F10', "Total Wallets Analyzed", str(kpi_metrics['total_wallets']), "4472C4")
    create_kpi_box(ws, 'G7', 'L10', "Total Profit (ETH)", f"{kpi_metrics['total_profit']:.4f}", "4472C4")
    
    create_kpi_box(ws, 'B11', 'F14', "Total Transactions", str(kpi_metrics['total_transactions']), "70AD47")
    create_kpi_box(ws, 'G11', 'L14', "Average Profit per TX", f"{kpi_metrics['avg_profit_per_tx']:.4f}", "70AD47")
    
    create_kpi_box(ws, 'B15', 'F18', "Most Profitable Wallet", kpi_metrics['top_wallet'][:8], "ED7D31")
    create_kpi_box(ws, 'G15', 'L18', "Top Wallet Profit", f"{kpi_metrics['top_wallet_profit']:.4f}", "ED7D31")
    
    # Add notes section
    ws.merge_cells('B20:L20')
    ws['B20'] = "ANALYSIS NOTES"
    ws['B20'].font = Font(size=14, bold=True)
    
    notes = [
        "• Data covers Ethereum Mainnet transactions for the specified address list.",
        "• Profit calculations include direct MEV gains minus transaction costs and bribes.",
        "• All profit values are reported in ETH equivalent at the time of transaction.",
        "• Performance and resimulation data relies on historical blockchain state.",
        f"• {kpi_metrics['total_wallets']} wallets were analyzed across {kpi_metrics['total_transactions']} transactions."
    ]
    
    for i, note in enumerate(notes):
        ws[f'B{22+i}'] = note
        ws[f'B{22+i}'].font = Font(size=11)
    
    # Navigation links
    ws.merge_cells('B28:L28')
    ws['B28'] = "NAVIGATION"
    ws['B28'].font = Font(size=14, bold=True)
    
    links = [
        ("View Leaderboard", "Leaderboard"),
    ]
    
    # Add links for each wallet sheet
    for address in all_results.keys():
        if all(r is not None for r in all_results[address][:4]):
            links.append((f"Wallet {address[:8]}", address[:8]))
    
    for i, (text, sheet) in enumerate(links):
        row = 30 + i
        cell = ws[f'B{row}']
        cell.value = text
        cell.hyperlink = f"#{sheet}!A1"
        cell.font = Font(color="0000FF", underline="single", size=12)
    
    # Set column widths
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    return ws

def calculate_kpi_metrics(all_results):
    """Calculate KPI metrics for the dashboard"""
    metrics = {
        'total_wallets': 0,
        'total_profit': 0.0,
        'total_transactions': 0,
        'avg_profit_per_tx': 0.0,
        'top_wallet': '',
        'top_wallet_profit': 0.0
    }
    
    valid_results = 0
    top_profit = 0
    
    for address, result in all_results.items():
        if not all(r is not None for r in result[:4]):
            continue
            
        metrics['total_wallets'] += 1
        combined_df, merged_df, libmev_df, final_df, latest_block = result
        
        if final_df is not None:
            wallet_profit = final_df['Profit'].sum()
            tx_count = len(final_df)
            
            metrics['total_profit'] += wallet_profit
            metrics['total_transactions'] += tx_count
            valid_results += 1
            
            # Check if this is the most profitable wallet
            if wallet_profit > top_profit:
                top_profit = wallet_profit
                metrics['top_wallet'] = address
                metrics['top_wallet_profit'] = wallet_profit
    
    # Calculate average profit per transaction
    if metrics['total_transactions'] > 0:
        metrics['avg_profit_per_tx'] = metrics['total_profit'] / metrics['total_transactions']
    
    return metrics

def create_kpi_box(ws, start_cell, end_cell, title, value, color):
    """Create a formatted KPI box in the dashboard"""
    ws.merge_cells(f'{start_cell}:{end_cell}')
    
    # Create styles
    title_font = Font(size=12, bold=True, color=color)
    value_font = Font(size=18, bold=True)
    
    # Add border to cell
    border = Border(
        left=Side(style='medium', color=color),
        right=Side(style='medium', color=color),
        top=Side(style='medium', color=color),
        bottom=Side(style='medium', color=color)
    )
    
    # Get cell reference
    cell = ws[start_cell]
    
    # Create a formatted string with title and value
    cell.value = f"{title}\n\n{value}"
    cell.font = title_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add slight fill
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    cell.fill = light_fill

def fetch_block_number(date_time):
    """Fetch the nearest block number for a given date and time"""
    timestamp = int(date_time.timestamp())
    
    # Use etherscan API to get the block number
    etherscan_api_key = os.environ.get("etherscan_api_key")
    url = f"https://api.etherscan.io/api?module=block&action=getblocknobytime&timestamp={timestamp}&closest=before&apikey={etherscan_api_key}"
    
    try:
        response = requests.get(url)
        data = response.json()
        if data["status"] == "1":
            return int(data["result"])
        logger.error(f"Failed to get block number: {data.get('message')}")
        # Fall back to a recent block
        return 19000000  # Use a reasonable fallback value
    except Exception as e:
        logger.error(f"Error fetching block number: {str(e)}")
        return 19000000  # Same fallback

def format_pl_report(address, df, output_xlsx, wb=None, ws=None):
    """
    Format PL report with professional styling
    Args:
        address: wallet address
        df: dataframe to write
        output_xlsx: output filename
        wb: optional workbook object (for multiple addresses)
        ws: optional worksheet object (for multiple addresses)
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.dimensions import ColumnDimension
    import math  # For isnan check

    def set_sheet_white(ws):
        """
        Set entire sheet to white background with no grid lines
        """
        from openpyxl.styles import PatternFill
        
        # Create white fill
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Apply to all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = white_fill
        
        # Turn off grid lines
        ws.sheet_view.showGridLines = False

    def convert_to_excel_friendly(value):
        if isinstance(value, list):
            return ', '.join(str(x) for x in value)
        elif isinstance(value, dict):
            if not value:
                return ''
            parts = []
            for k, v in value.items():
                if isinstance(v, dict):
                    nested_parts = [f"{sub_k}: {sub_v}" for sub_k, sub_v in v.items()]
                    parts.append(f"{k}: {{{', '.join(nested_parts)}}}")
                else:
                    parts.append(f"{k}: {v}")
            return '{' + ', '.join(parts) + '}'
        return value

    def clean_illegal_excel_chars(value):
        """Clean string values to remove illegal Excel characters"""
        if isinstance(value, str):
            # Handle common resim cases
            resim_replacements = {
                'Revertᴿ': 'Revert:R',
                'Revert:Rˡ': 'Revert:R',
                'ˡ': 'l',
                'ᴿ': 'R',
                '\u0000': '',  # Null character
            }

            for old, new in resim_replacements.items():
                value = value.replace(old, new)

            # Remove any other control characters
            value = ''.join(char for char in value if ord(char) >= 32 or char in '\n\r\t')

        return value

    def safe_numeric(value):
        """
        Safely convert any numeric type to float.
        """
        if value is None:
            return None

        try:
            result = float(value)
            return result
        except (ValueError, TypeError):
            return None

    def get_column_letter_by_header(worksheet, header_text, header_row=9):
        """
        Find the column letter based on the header text in the specified row.
        
        Args:
            worksheet: The openpyxl worksheet
            header_text: The text to search for in the header row
            header_row: The row number containing headers (default 9)
        
        Returns:
            str: Column letter (e.g., 'A', 'B', 'AA') or None if not found
        """
        from openpyxl.utils import get_column_letter
        
        for cell in worksheet[header_row]:
            if cell.value == header_text:
                return get_column_letter(cell.column)
        return None

    # Clean the dataframe
    for column in df.columns:
        df[column] = df[column].apply(lambda x:
            clean_illegal_excel_chars(
                f"Error: {x['error']['message']}" if isinstance(x, dict) and 'error' in x
                else convert_to_excel_friendly(x)
            )
        )

    # Initial Excel setup
    if wb is None or ws is None:
        df.to_excel(output_xlsx, sheet_name='PL Report', startrow=8, startcol=1, index=False)
        wb = load_workbook(output_xlsx)
        ws = wb['PL Report']
        set_sheet_white(ws)
    else:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 9):
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx, column=c_idx, value=value)
        set_sheet_white(ws)

    # Handle exception for 0x75efe to replace my profit calcs with those from libmev - for now at least
    if address == "0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D":
        # Get column letters based on headers
        profit_col = get_column_letter_by_header(ws, 'Profit')
        profit_eth_col = get_column_letter_by_header(ws, 'profit_eth')
        margin_col = get_column_letter_by_header(ws, 'Margin %')
        profit_margin_col = get_column_letter_by_header(ws, 'profit_margin')
        tx_cost_col = get_column_letter_by_header(ws, 'Tx Cost (ETH)')
        extractable_col = get_column_letter_by_header(ws, 'Extractable Value')
        rank_col = get_column_letter_by_header(ws, 'Profit Rank')
        
        # Get the last row with data
        last_row = ws.max_row
        
        # 1. Replace Profit values with profit_eth
        if profit_col and profit_eth_col:
            for row in range(10, last_row + 1):
                profit_eth_cell = ws[f'{profit_eth_col}{row}']
                profit_cell = ws[f'{profit_col}{row}']
                
                # Convert None, empty strings, or NaN to 0
                if profit_eth_cell.value in (None, '', 'nan') or (isinstance(profit_eth_cell.value, float) and math.isnan(profit_eth_cell.value)):
                    profit_value = 0.0
                else:
                    profit_value = safe_numeric(profit_eth_cell.value)
                    if profit_value is None:  # If conversion failed
                        profit_value = 0.0
                
                profit_cell.value = profit_value
                profit_cell.number_format = '0.0000'  # 4 decimal places

        # 2. Replace Margin % with profit_margin and format as percentage
        if margin_col and profit_margin_col:
            for row in range(10, last_row + 1):
                margin_cell = ws[f'{margin_col}{row}']
                profit_margin_cell = ws[f'{profit_margin_col}{row}']
                margin_value = safe_numeric(profit_margin_cell.value)
                if margin_value is not None:
                    margin_cell.value = margin_value
                    margin_cell.number_format = '0.0%'  # Percentage with 1 decimal

        # 3. Recalculate Extractable Value as Profit + Tx Cost
        if tx_cost_col and profit_col and extractable_col:
            for row in range(10, last_row + 1):
                tx_cost_cell = ws[f'{tx_cost_col}{row}']
                profit_cell = ws[f'{profit_col}{row}']
                extractable_cell = ws[f'{extractable_col}{row}']
                
                tx_cost = safe_numeric(tx_cost_cell.value)
                profit = safe_numeric(profit_cell.value)
                if tx_cost is not None and profit is not None:
                    extractable_cell.value = tx_cost + profit
                    extractable_cell.number_format = '0.0000'

        # 4. Sort by Profit descending and update ranks
        if profit_col and rank_col:
            # Store all row data with formatting
            all_rows = []
            for row in range(10, last_row + 1):
                row_data = []
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    # Store cell properties
                    font_props = {}
                    if cell.font:
                        font_props = {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'color': cell.font.color,
                            'underline': cell.font.underline,
                            'strike': cell.font.strike,
                            'vertAlign': cell.font.vertAlign
                        }
                    
                    alignment_props = {}
                    if cell.alignment:
                        alignment_props = {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical,
                            'wrap_text': cell.alignment.wrap_text,
                            'shrink_to_fit': cell.alignment.shrink_to_fit,
                            'indent': cell.alignment.indent
                        }
                    
                    row_data.append({
                        'value': cell.value if cell.value is not None else 0.0,
                        'number_format': cell.number_format,
                        'font_props': font_props,
                        'alignment_props': alignment_props,
                        'hyperlink': cell.hyperlink
                    })
                all_rows.append(row_data)
            
            # Sort rows by profit value
            profit_idx = ord(profit_col.upper()) - ord('B')
            sorted_rows = sorted(
                all_rows,
                key=lambda x: (safe_numeric(x[profit_idx]['value']) or 0.0),
                reverse=True
            )
            
            # Write back sorted data
            from openpyxl.styles import Font, Alignment
            for i, row_data in enumerate(sorted_rows):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=i+10, column=j+2)
                    cell.value = cell_data['value'] if cell_data['value'] != 0.0 else None
                    cell.number_format = cell_data['number_format']
                    
                    if cell_data['font_props']:
                        cell.font = Font(**cell_data['font_props'])
                    
                    if cell_data['alignment_props']:
                        cell.alignment = Alignment(**cell_data['alignment_props'])
                    
                    cell.hyperlink = cell_data['hyperlink']
            
            # Update ranks
            for row in range(10, last_row + 1):
                rank_cell = ws[f'{rank_col}{row}']
                rank_cell.value = row - 9  # Simple 1-based ranking
    
    header_font = Font(name='Calibri', size=14, bold=True)
    normal_font = Font(name='Calibri', size=12)
    mono_font = Font(name='Consolas', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set initial properties
    ws.sheet_view.zoomScale = 140
    ws['C2'] = address
    
    # Get the entity name if available
    entity = address_to_entity.get(address.lower(), None)
    if entity:
        ws['C3'] = f"Entity: {entity}"
        ws['C3'].font = Font(size=14, italic=True)
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

    # 1. Add Bundle Tx Details columns
    def get_max_bundle_transactions(df, count_column_name):
        try:
            return int(df[count_column_name].max())
        except KeyError:
            return 0

    def add_bundle_tx_details(ws, max_tx, start_col_letter='AX', header_font=None, thin_border=None):
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter, column_index_from_string

        start_col = column_index_from_string(start_col_letter)
        header_row = 9
        header_cell = ws.cell(row=header_row, column=start_col)
        header_cell.value = "Bundle Tx Details"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if header_font:
            header_cell.font = header_font
        if thin_border:
            header_cell.border = thin_border
        
        if max_tx > 1:
            end_col = start_col + max_tx - 1
            ws.merge_cells(
                start_row=header_row, 
                start_column=start_col, 
                end_row=header_row, 
                end_column=end_col
            )

        for i in range(max_tx):
            col = start_col + i
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions[col_letter].alignment = Alignment(horizontal='center')

    def process_transaction_details(ws, df, tx_column_name, searcher_column_name, start_col_letter='AX'):
        from openpyxl.styles import Font, Alignment
        from openpyxl.comments import Comment
        from openpyxl.utils import column_index_from_string

        start_col = column_index_from_string(start_col_letter)

        blue_font = Font(color="0000FF", underline="single")
        green_font = Font(color="008000", underline="single")
        center_align = Alignment(horizontal='center', vertical='center')

        # Set column width for the range
        max_tx_count = 0
        for row in range(10, ws.max_row + 1):
            tx_value = df.iloc[row-10][tx_column_name]
            if pd.notna(tx_value):
                tx_count = len(str(tx_value).split(','))
                max_tx_count = max(max_tx_count, tx_count)

        column_width = 3
        
        # Apply column width to all potentially used columns
        for i in range(max_tx_count):
            col_letter = get_column_letter(start_col + i)
            ws.column_dimensions[col_letter].width = column_width
        
        for row in range(10, ws.max_row + 1):
            tx_value = df.iloc[row-10][tx_column_name]
            searcher_txs = df.iloc[row-10][searcher_column_name]

            if pd.notna(tx_value):
                txs = [tx.strip() for tx in str(tx_value).split(',')]
                searcher_set = set(tx.strip() for tx in str(searcher_txs).split(',')) if pd.notna(searcher_txs) else set()

                for i, tx in enumerate(txs):
                    detail_cell = ws.cell(row=row, column=start_col + i)
                    detail_cell.value = f"Tx{i+1}"
                    detail_cell.hyperlink = f'https://etherscan.io/tx/{tx}'
                    detail_cell.alignment = Alignment(horizontal='center')
                    detail_cell.border = thin_border

                    if tx in searcher_set:
                        detail_cell.font = blue_font
                    else:
                        detail_cell.font = green_font

                    comment = Comment(f"Transaction Hash:\n{tx}", "System")
                    detail_cell.comment = comment

    max_tx = get_max_bundle_transactions(df, "# Txs in Bundle")

    if max_tx > 0:
        add_bundle_tx_details(ws, max_tx, start_col_letter='AX', header_font=header_font, thin_border=thin_border)
        process_transaction_details(ws, df, "Txs in Bundle", "searcher_txs", start_col_letter='AX')

    # 2. Add additional columns
    ws['AT9'] = 'Tenderly'
    ws['AU9'] = 'Eigenphi tx'
    ws['AV9'] = 'Eigenphi block'
    ws['AW9'] = 'libMEV'

    # 3. Move column X to AW (if needed)
    def move_column_to_end(ws, from_col_letter, to_col_letter):
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
        from openpyxl.utils import column_index_from_string
        
        from_col = column_index_from_string(from_col_letter)
        to_col = column_index_from_string(to_col_letter)
        
        # Store column data
        col_data = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=from_col)
            if not isinstance(cell, MergedCell):
                # Get border properties
                border_kwargs = {}
                if cell.border:
                    for side in ['left', 'right', 'top', 'bottom']:
                        side_obj = getattr(cell.border, side)
                        if side_obj:
                            border_kwargs[side] = Side(
                                style=side_obj.style,
                                color=side_obj.color
                            )
                
                # Get fill properties
                fill_kwargs = None
                if cell.fill and cell.fill.fill_type != None:
                    fill_kwargs = {
                        'fill_type': cell.fill.fill_type,
                        'start_color': cell.fill.start_color,
                        'end_color': cell.fill.end_color
                    }
                
                col_data.append({
                    'value': cell.value,
                    'hyperlink': cell.hyperlink,
                    'font_kwargs': {
                        'name': cell.font.name,
                        'size': cell.font.size,
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'color': cell.font.color,
                        'underline': cell.font.underline,
                        'strike': cell.font.strike,
                        'vertAlign': cell.font.vertAlign
                    } if cell.font else None,
                    'alignment_kwargs': {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrap_text': cell.alignment.wrap_text,
                        'shrink_to_fit': cell.alignment.shrink_to_fit,
                        'indent': cell.alignment.indent
                    } if cell.alignment else None,
                    'border_kwargs': border_kwargs,
                    'fill_kwargs': fill_kwargs,
                    'number_format': cell.number_format,
                    'protection': cell.protection.locked if cell.protection else None,
                    'comment': cell.comment
                })
            else:
                col_data.append(None)
        
        # Shift columns left to fill gap
        for row in range(1, ws.max_row + 1):
            for col in range(from_col, to_col):
                source_cell = ws.cell(row=row, column=col + 1)
                target_cell = ws.cell(row=row, column=col)
                if not isinstance(target_cell, MergedCell) and not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value
                    target_cell.hyperlink = source_cell.hyperlink
                    
                    if source_cell.font:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            vertAlign=source_cell.font.vertAlign
                        )
                    
                    if source_cell.alignment:
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text,
                            shrink_to_fit=source_cell.alignment.shrink_to_fit,
                            indent=source_cell.alignment.indent
                        )
                    
                    if source_cell.border:
                        border_kwargs = {}
                        for side in ['left', 'right', 'top', 'bottom']:
                            side_obj = getattr(source_cell.border, side)
                            if side_obj:
                                border_kwargs[side] = Side(
                                    style=side_obj.style,
                                    color=side_obj.color
                                )
                        target_cell.border = Border(**border_kwargs)
                    
                    if source_cell.fill and source_cell.fill.fill_type != None:
                        target_cell.fill = PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                    
                    target_cell.number_format = source_cell.number_format
                    if source_cell.protection:
                        target_cell.protection = Protection(locked=source_cell.protection.locked)
                    target_cell.comment = source_cell.comment
    
        # Write stored data to new position
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=to_col)
            if not isinstance(cell, MergedCell) and col_data[row - 1]:
                data = col_data[row - 1]
                cell.value = data['value']
                cell.hyperlink = data['hyperlink']
                if data['font_kwargs']:
                    cell.font = Font(**data['font_kwargs'])
                if data['alignment_kwargs']:
                    cell.alignment = Alignment(**data['alignment_kwargs'])
                if data['border_kwargs']:
                    cell.border = Border(**data['border_kwargs'])
                if data['fill_kwargs']:
                    cell.fill = PatternFill(**data['fill_kwargs'])
                cell.number_format = data['number_format']
                if data['protection']:
                    cell.protection = Protection(locked=data['protection'])
                cell.comment = data['comment']

    move_column_to_end(ws, 'X', 'AW')

    # 4. Apply all formatting
    # Set background white for all cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = white_fill

    # Format column headers
    for row in ws['B9:' + get_column_letter(ws.max_column) + '9']:
        for cell in row:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Update filter range
    ws.auto_filter.ref = f'B9:{get_column_letter(ws.max_column)}{ws.max_row}'

    # Format data cells
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            if cell.column < 2:
                continue

            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')

            col_letter = get_column_letter(cell.column)
            header_cell = ws[f'{col_letter}9']

            if header_cell.value in ['block #', 'nonce', 'tx Index', 'Profit Rank', 'Resim', 
                                   'Resim @ txIx 0', 'Resim @ txIx -1', 'Tenderly', 
                                   'Eigenphi tx', 'Eigenphi block', 'libMEV',
                                   '# Txs in Bundle', 'Bundle Tx Details', 'searcher_txs_count', 'tokens_count']:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if header_cell.value in ['from', 'to', 'methodId']:
                cell.font = mono_font

            elif header_cell.value in ['tx Hash', 'tx Hash']:
                cell.font = mono_font
                if cell.value:
                    full_hash = str(cell.value)
                    cell.value = str(cell.value)[:10] + '...'
                    cell.hyperlink = f'https://etherscan.io/tx/{full_hash}'
                    cell.font = Font(name='Consolas', size=12, color='0000FF', underline='single')

            elif header_cell.value == 'block #':
                cell.font = mono_font
                if cell.value:
                    block_no = str(cell.value)
                    cell.hyperlink = f'https://etherscan.io/txs?block={block_no}'
                    cell.font = Font(name='Consolas', size=12, color='0000FF', underline='single')

            elif header_cell.value == 'Tenderly':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "tx"
                cell.hyperlink = f'https://dashboard.tenderly.co/tx/mainnet/{full_hash}'

            elif header_cell.value == 'Eigenphi tx':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "tx"
                cell.hyperlink = f'https://eigenphi.io/mev/ethereum/tx/{full_hash}'

            elif header_cell.value == 'Eigenphi block':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "block"
                cell.hyperlink = f'https://eigenphi.io/mev/eigentx/{full_hash}?tab=block'

            elif header_cell.value == 'libMEV':
                cell.font = Font(color='0000FF', underline='single')
                cell.value = "block"
                cell.hyperlink = f'https://libmev.com/blocks/{block_no}'

            elif header_cell.value in ['Resim @ txIx 0', 'Resim @ txIx -1']:
                if cell.value == 'OK':
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

            elif header_cell.value == 'dateTime':
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Format summary section
    def format_summary_headers(cell):
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(size=14, bold=True)

    def format_summary_values(cell):
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=14, bold=False)

    # Add and format summary headers
    summary_headers = {
        'C4': 'Start Block:', 'C5': 'End Block:', 'C6': 'Total Transactions:', 'C7': 'Total Profit:',
        'G4': 'Start Balance:', 'G5': 'End Balance:', 'G6': 'Withdrawals:', 'G7': 'Total Margin:'
    }

    for pos, text in summary_headers.items():
        ws[pos] = text
        format_summary_headers(ws[pos])

    # Add formulas
    formulas = {
        'D4': f'=MIN(B10:B{len(df)+9})',
        'D5': f'=MAX(B10:B{len(df)+9})',
        'D6': f'=COUNTA(B10:B{len(df)+9})',
        'D7': f'=SUM(H5-H4+H6)',
        'H4': f'=VLOOKUP(D4,B10:H{len(df)+9},7,FALSE)',
        'H5': f'=VLOOKUP(D5,B10:H{len(df)+9},7,FALSE)',
        'H6': f'=SUMIF(I10:I{len(df)+9},"Withdrawal",Q10:Q{len(df)+9})',
        'H7': f'=TEXT(ROUND((1-(SUM(N10:O{len(df)+9})/SUM(P10:P{len(df)+9}))), 5), "0.0%")'
    }

    for pos, formula in formulas.items():
        ws[pos] = formula
        format_summary_values(ws[pos])

    if address == "0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D":
            special_formulas = {
                'D7': f'=ROUND(SUM($P$10:$P${len(df)+9})+$H$6,3)',  # Added $ for absolute references
            }
            for pos, formula in special_formulas.items():
                ws[pos] = formula
                format_summary_values(ws[pos])

    # After the header formatting loop where you set font, alignment, and border
    for row in ws['B9:' + get_column_letter(ws.max_column) + '9']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set row height to auto
    ws.row_dimensions[9].height = None  # This triggers auto-height calculation

    # Format address header
    ws.merge_cells('C2:G2')
    cell = ws['C2']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=16, bold=True)

    # 5. Apply column widths
    width_groups = {
        20: ['A', 'D'],
        17.5: ['C', 'G'],
        15: ['B'],
        10: ['H', 'O', 'P', 'U', 'V', 'X', 'AS', 'AT'],
        8.5: ['E', 'F', 'S', 'AW'],
        7.5: ['Q', 'S', 'T', 'AU', 'AV'],
        5: ['AX', 'AY', 'AZ']
    }

    for width, columns in width_groups.items():
        for col in columns:
            try:
                ws.column_dimensions[col].width = width
            except:
                pass

    # 6. Apply grouping and hiding (LAST)
    def get_columns_in_range(start_col, end_col):
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

    # Group and hide columns
    for col_letter in get_columns_in_range('I', 'P'):
        ws.column_dimensions[col_letter].outline_level = 1
        ws.column_dimensions[col_letter].hidden = True

    for col_letter in get_columns_in_range('W', 'AR'):
        ws.column_dimensions[col_letter].outline_level = 1
        ws.column_dimensions[col_letter].hidden = True

    # Final formatting tweaks
    # Alternate colors only from column B
    for row in range(10, ws.max_row + 1):
        if (row - 9) % 2 == 0:
            for cell in ws[row]:
                if 2 <= cell.column <= len(df.columns) + 2:  # Only B onwards
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # Add profit column color scale
    profit_col = get_column_letter(df.columns.get_loc('Profit') + 2)  # +2 for startcol offset
    ws.conditional_formatting.add(
        f'{profit_col}10:{profit_col}{ws.max_row}',
        ColorScaleRule(
            start_type='min',
            start_color='FF0000',
            mid_type='percentile', 
            mid_value=50,
            mid_color='FFFF00',
            end_type='max',
            end_color='00FF00'
        )
    )

    ws.merge_cells('AT9:AU9')
    ws['AT9'] = "Eigenphi"

    def format_bundle_tx_details(ws):
        """
        Format all columns under 'Bundle Tx Details' with consistent styling
        Groups columns after the first 5
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font
    
        # Find the Bundle Tx Details header (should be in row 9)
        start_col = None
        end_col = None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=9, column=col)
            if cell.value == "Bundle Tx Details":
                if start_col is None:
                    start_col = col
                end_col = col
    
        if start_col is None:
            logger.warning("Bundle Tx Details header not found")
            return
    
        # Get last row with data
        last_row = ws.max_row
    
        # Create style objects
        alignment = Alignment(horizontal='center', vertical='center')
        font = Font(size=8, color="0000FF", underline="single")
        column_width = 2.5
    
        # Count how many columns are in the Bundle Tx Details section
        bundle_cols = []
        for col in range(start_col, ws.max_column + 1):
            # Check if we've moved past the Bundle Tx Details section
            header_value = ws.cell(row=9, column=col).value
            if header_value and header_value != "Bundle Tx Details":
                break
            bundle_cols.append(col)
    
        total_bundle_cols = len(bundle_cols)
    
        # Format all columns
        for col in bundle_cols:
            col_letter = get_column_letter(col)
            
            # Set column width
            ws.column_dimensions[col_letter].width = column_width
            
            # Apply formatting to cells
            for row in range(10, last_row + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = alignment
                    cell.font = font
                except:
                    pass
        
            # Group columns after the first 5
            if col >= start_col + 5:  # If this is the 6th or later column
                ws.column_dimensions[col_letter].outline_level = 1
                ws.column_dimensions[col_letter].hidden = True  # Initially hidden
    
        # If we have more than 5 columns, make sure outlining is enabled
        if total_bundle_cols > 5:
            ws.sheet_view.showOutlineSymbols = True

    format_bundle_tx_details(ws)

    def add_bribe_percentage(ws):
        """
        Insert a new 'Bribe %' column after 'Margin %' and group them
        Uses Excel formulas to calculate Bribe % as (1 - Margin %)
        Completely matches formatting from Margin % (percentage format, borders, shading)
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        
        # Find the 'Margin %' column
        margin_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=9, column=col).value == "Margin %":
                margin_col = col
                break
                
        if margin_col is None:
            logger.warning("Margin % column not found")
            return
            
        # Get the last row with data
        last_row = ws.max_row
        
        # Insert a new column after Margin %
        ws.insert_cols(idx=margin_col + 1)
        
        # Get formatting from margin header cell
        margin_header = ws.cell(row=9, column=margin_col)
        
        # Set the header for the new column (create new style objects)
        bribe_header = ws.cell(row=9, column=margin_col + 1)
        bribe_header.value = "Bribe %"
        
        # Create new Font object
        if margin_header.font:
            bribe_header.font = Font(
                name=margin_header.font.name,
                size=margin_header.font.size,
                bold=margin_header.font.bold,
                italic=margin_header.font.italic,
                color=margin_header.font.color
            )
        
        # Create new Alignment object
        bribe_header.alignment = Alignment(horizontal='center', vertical='center')
        
        # Create new Border object if needed
        if margin_header.border:
            border_sides = {}
            for side in ['left', 'right', 'top', 'bottom']:
                side_obj = getattr(margin_header.border, side)
                if side_obj and side_obj.style:
                    border_sides[side] = Side(style=side_obj.style, color=side_obj.color)
            
            if border_sides:
                bribe_header.border = Border(**border_sides)
        
        # Copy Fill (background shading)
        if margin_header.fill and margin_header.fill.fill_type != 'none':
            bribe_header.fill = PatternFill(
                fill_type=margin_header.fill.fill_type,
                start_color=margin_header.fill.start_color,
                end_color=margin_header.fill.end_color
            )
        
        # Get margin column letter for formulas
        margin_col_letter = get_column_letter(margin_col)
        
        # Add formula for each row: Bribe % = 1 - Margin %
        for row in range(10, last_row + 1):
            margin_cell = ws.cell(row=row, column=margin_col)
            bribe_cell = ws.cell(row=row, column=margin_col + 1)
            
            # Set formula: =1-[Margin %]
            bribe_cell.value = f"=1-{margin_col_letter}{row}"
            
            # Ensure percentage formatting
            if "%" in margin_cell.number_format:
                bribe_cell.number_format = margin_cell.number_format
            else:
                # If margin is not formatted as percentage, apply a percentage format
                bribe_cell.number_format = "0.0%"
            
            # Create new Font object if needed
            if margin_cell.font:
                bribe_cell.font = Font(
                    name=margin_cell.font.name,
                    size=margin_cell.font.size,
                    bold=margin_cell.font.bold,
                    italic=margin_cell.font.italic,
                    color=margin_cell.font.color
                )
            
            # Create new Alignment object
            if margin_cell.alignment:
                bribe_cell.alignment = Alignment(
                    horizontal=margin_cell.alignment.horizontal,
                    vertical=margin_cell.alignment.vertical,
                    wrap_text=margin_cell.alignment.wrap_text
                )
                
            # Create Border object
            if margin_cell.border:
                border_sides = {}
                for side in ['left', 'right', 'top', 'bottom']:
                    side_obj = getattr(margin_cell.border, side)
                    if side_obj and side_obj.style:
                        border_sides[side] = Side(style=side_obj.style, color=side_obj.color)
                
                if border_sides:
                    bribe_cell.border = Border(**border_sides)
            
            # Copy Fill (background shading)
            if margin_cell.fill and margin_cell.fill.fill_type != 'none':
                bribe_cell.fill = PatternFill(
                    fill_type=margin_cell.fill.fill_type,
                    start_color=margin_cell.fill.start_color,
                    end_color=margin_cell.fill.end_color
                )
        
        # Group the columns and hide Margin %
        ws.column_dimensions.group(margin_col_letter, margin_col_letter, hidden=True)
        
        # Set column width of Bribe % to match Margin % width
        bribe_col_letter = get_column_letter(margin_col + 1)
        ws.column_dimensions[bribe_col_letter].width = ws.column_dimensions[margin_col_letter].width

    add_bribe_percentage(ws)

    # Set row height
    ws.row_dimensions[10].height = 20

    # Set worksheet title
    ws.title = address[:8]

    # Save workbook
    wb.save(output_xlsx)

def main():
    """Main execution function"""
    # Parse command line arguments
    args = parse_arguments()

    # Configure batch size
    batch_size = args.batch_size if args.batch_size else MAX_BATCH_SIZE


    # Get addresses
    if args.addresses:
        address_list = [addr.strip() for addr in args.addresses.split(',')]
    else:
        address_list = [address for group in entity_groups.values() for address in group]

    if not address_list or not address_list[0]:
        logger.error("No addresses provided. Use --addresses or set address_list in .env")
        return

    # Handle time period selection
    if args.time_period == 'custom':
        if not args.start_date:
            logger.error("Error: --start_date is required when using --time_period custom")
            return

        try:
            start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
            end_date = datetime.strptime(args.end_date, '%Y-%m-%d')

            if start_date > end_date:
                logger.error("Error: start_date must be before end_date")
                return

            # Get block for start date
            logger.info(f"Getting block number for date {args.start_date}")
            start_block = fetch_block_number(start_date)

            # Use manual time_data structure for custom range
            time_data = {
                'custom': {
                    'Ethereum': {
                        'block_number': start_block,
                        'datetime': start_date.strftime('%Y-%m-%d %H:%M:%S'),
                        'timestamp': int(start_date.timestamp())
                    }
                }
            }
            time_period = 'custom'
        except ValueError:
            logger.error("Error: Invalid date format. Use YYYY-MM-DD")
            return
    else:
        # Get block data using the specified time period
        time_period = args.time_period
        logger.info(f"Getting block data for time period: {time_period}")
 
        from block_calculation_engine import main as block_engine_main

        time_data = block_engine_main(["Ethereum"], [time_period])

    start_block = int(time_data[time_period]['Ethereum']['block_number'])
    logger.info(f"Starting from block: {start_block}")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Process all addresses
    first_run = True
    output_xlsx = args.output if args.output else None
    all_results = {}

    for addr in address_list:
        logger.info(f"Processing address: {addr}")
        
        # Calculate adjusted start block - for most addresses, use the standard block
        if addr in ["0x1b9FcB24c533839dC847235bd8Eb80E37EC42f85"]:
            start_block_adj = start_block
        elif addr in ["0x0BdE59981FDEaC219Ce9E618d27F193438Bff786"]:
            start_block_adj = start_block
        elif addr in ["0xe75eD6F453c602Bd696cE27AF11565eDc9b46B0D"]:
            start_block_adj = start_block
        else:
            start_block_adj = start_block
        
        # Process address
        result = process_address(addr, start_block_adj, batch_size)
        all_results[addr] = result
        
        # Set output filename based on first successful result
        if first_run and result[4] is not None:
            try:
                # Convert to int if it's not already
                latest_block = int(result[4]) if isinstance(result[4], str) else result[4]
                if latest_block > 0:
                    if not output_xlsx:
                        output_xlsx = f'MEV_daily_multiple_addresses_{latest_block}_P&L.xlsx'
                    first_run = False
            except (ValueError, TypeError):
                logger.warning(f"Couldn't use block number {result[4]} for filename")
        
        # Create worksheet for this address if processing was successful
        if result[3] is not None:  # Check if final_df exists
            combined_df, merged_df, libmev_df, final_df, latest_block = result
            ws = wb.create_sheet(title=addr[:8])
            format_pl_report(addr, final_df, output_xlsx, wb, ws)
            logger.info(f"Added worksheet for {addr[:8]}")
    
    # Add summary/leaderboard sheet
    create_summary_leaderboard(all_results, wb)
    
    # Add dashboard sheet (if at least one address was processed successfully)
    if any(result[3] is not None for result in all_results.values()):
        create_dashboard_sheet(wb, all_results)
    
    # Save the workbook
    if not output_xlsx:
        output_xlsx = f'MEV_analysis_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb.save(output_xlsx)
    logger.info(f"Analysis complete. Results saved to {output_xlsx}")

if __name__ == "__main__":
    main()
