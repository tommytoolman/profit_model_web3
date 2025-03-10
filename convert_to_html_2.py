"""
Complete HTML Export Module with Nested Tab Navigation and Landscape Support
"""
import os
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Union

from metadata import entity_groups

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def get_entity_group(sheet_name: str) -> str:
    """
    Determine which entity group a sheet belongs to based on the first 8 characters.
    
    Args:
        sheet_name: The name of the worksheet (typically an address)
        
    Returns:
        The primary group this address belongs to
    """
    # Extract the first 8 characters of the sheet name
    sheet_prefix = sheet_name[:8].lower() if len(sheet_name) >= 8 else sheet_name.lower()
    
    # Check if this matches any primary entity or sub-entity
    for primary_entity, sub_entities in entity_groups.items():
        # Check if sheet matches the primary entity
        primary_prefix = primary_entity[:8].lower() if len(primary_entity) >= 8 else primary_entity.lower()
        if sheet_prefix == primary_prefix:
            return primary_entity
        
        # Check if sheet matches any sub-entity
        for sub_entity in sub_entities:
            sub_prefix = sub_entity[:8].lower() if len(sub_entity) >= 8 else sub_entity.lower()
            if sheet_prefix == sub_prefix:
                return primary_entity
    
    # Default group if not found
    return "Other" 


def get_friendly_name(address: str, max_length: int = 8) -> str:
    """
    Generate a shorter, more readable name for an address.
    
    Args:
        address: The full blockchain address
        max_length: Maximum characters to show
        
    Returns:
        Shortened address (e.g. '0x1b9F...')
    """
    if not address:
        return "Unknown"
    
    if len(address) <= max_length:
        return address
    
    return f"{address[:max_length]}..."


def get_cell_value(cell: Cell, is_block_number: bool = False) -> str:
    """Extract and format cell value, handling different types."""
    if cell.value is None:
        return ""
    
    # Handle numeric values with specific formatting
    if is_block_number and isinstance(cell.value, (int, float)):
        # Plain integer format for block numbers
        return f"{int(cell.value)}"
    elif cell.number_format.endswith('%') and isinstance(cell.value, (int, float)):
        return f"{cell.value:.1%}"
    elif isinstance(cell.value, (int, float)):
        return f"{cell.value:,.4f}" if cell.value % 1 != 0 else f"{cell.value:,.0f}"
    
    return str(cell.value)


def get_cell_hyperlink(cell: Cell) -> Optional[str]:
    """Extract hyperlink from cell if present."""
    return cell.hyperlink.target if cell.hyperlink else None


def get_cell_style(cell: Cell) -> Dict:
    """Extract cell styling information."""
    style = {}
    
    # Get text color
    if cell.font and cell.font.color:
        if cell.font.color.rgb:
            rgb = cell.font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:  # ARGB format
                style['color'] = f"#{rgb[2:]}"  # Remove alpha channel
    
    # Get bold/italic status
    if cell.font:
        if cell.font.bold:
            style['font-weight'] = 'bold'
        if cell.font.italic:
            style['font-style'] = 'italic'
        if cell.font.underline:
            style['text-decoration'] = 'underline'
    
    # Get background color
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        rgb = cell.fill.start_color.rgb
        if isinstance(rgb, str) and len(rgb) == 8:  # ARGB format
            if rgb != 'FF000000':  # Not default black
                style['background-color'] = f"#{rgb[2:]}"  # Remove alpha channel
    
    # Get alignment
    if cell.alignment:
        if cell.alignment.horizontal:
            style['text-align'] = cell.alignment.horizontal
        if cell.alignment.vertical:
            style['vertical-align'] = cell.alignment.vertical
    
    return style


def extract_worksheet_data(ws: Worksheet, keep_columns: List[str] = None, 
                          max_rows_per_sheet: int = None) -> Tuple[Dict, List[Dict]]:
    """
    Extract data from Excel worksheet into structured format for HTML export.
    
    Args:
        ws: OpenPyXL worksheet object
        keep_columns: List of column headers to include (if None, include all)
        max_rows_per_sheet: Maximum number of rows to extract per sheet (for size control)
    
    Returns:
        Tuple containing (summary_data, main_table_data)
    """
    # Skip Dashboard and Leaderboard tabs
    if ws.title in ['Dashboard', 'Leaderboard']:
        return None, None
    
    # First, extract the summary data (C2:H7)
    summary_data = {
        'address': ws['C2'].value,
        'summary_rows': []
    }
    
    # Extract summary rows (4-7)
    for row in range(4, 8):
        # Special handling for Start Block and End Block - remove commas
        is_block_cell = ws[f'C{row}'].value in ['Start Block:', 'End Block:']
        
        summary_data['summary_rows'].append({
            'label': get_cell_value(ws[f'C{row}']),
            'value': get_cell_value(ws[f'D{row}'], is_block_cell),
            'label_style': get_cell_style(ws[f'C{row}']),
            'value_style': get_cell_style(ws[f'D{row}']),
            'extra_label': get_cell_value(ws[f'G{row}']),
            'extra_value': get_cell_value(ws[f'H{row}']),
            'extra_label_style': get_cell_style(ws[f'G{row}']),
            'extra_value_style': get_cell_style(ws[f'H{row}'])
        })
    
    # Now extract the header row (row 9)
    headers = []
    header_indices = {}  # Map header names to column indices
    excluded_cols = ['dateTime', 'Extractable Value', 'Total Tip']
    
    # First collect all headers and their positions
    for col_idx in range(2, ws.max_column + 1):  # Start from column B (index 2)
        cell = ws.cell(row=9, column=col_idx)
        header_text = get_cell_value(cell)
        if header_text:
            # Skip excluded columns
            if header_text in excluded_cols:
                continue
                
            # Rename 'Balance (ETH + WETH)' to 'Balance'
            if header_text == 'Balance (ETH + WETH)':
                header_text = 'Balance'
                
            header_indices[header_text] = col_idx
    
    # Process columns in specific order to move 'miner' to the end
    if keep_columns:
        # First, remove miner from the list
        miner_col = None
        new_columns = []
        for col in keep_columns:
            if col != 'miner' and col not in excluded_cols:
                new_columns.append(col)
            elif col == 'miner':
                miner_col = col
        
        # Then add it back at the end
        if miner_col:
            new_columns.append(miner_col)
        
        # Now process the columns in this order
        for header_name in new_columns:
            # Handle the renamed column
            lookup_name = 'Balance (ETH + WETH)' if header_name == 'Balance' else header_name
            if lookup_name in header_indices:
                col_idx = header_indices[lookup_name]
                cell = ws.cell(row=9, column=col_idx)
                headers.append({
                    'text': header_name,  # Use the new name
                    'style': get_cell_style(cell),
                    'column': col_idx
                })
    else:
        # If no columns specified, keep all in original order (except excluded ones)
        # but still move miner to the end
        miner_col = None
        miner_data = None
        
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=9, column=col_idx)
            header_text = get_cell_value(cell)
            
            if header_text and header_text not in excluded_cols:
                if header_text == 'miner':
                    miner_col = col_idx
                    miner_data = {
                        'text': 'miner',
                        'style': get_cell_style(cell),
                        'column': col_idx
                    }
                else:
                    # Rename 'Balance (ETH + WETH)' to 'Balance'
                    if header_text == 'Balance (ETH + WETH)':
                        header_text = 'Balance'
                    
                    headers.append({
                        'text': header_text,
                        'style': get_cell_style(cell),
                        'column': col_idx
                    })
        
        # Add miner at the end if it exists
        if miner_data:
            headers.append(miner_data)
    
    # Extract data rows
    data_rows = []
    
    # Determine the range of rows to extract
    max_row = ws.max_row
    if max_rows_per_sheet and max_rows_per_sheet < (max_row - 9):
        max_row = 9 + max_rows_per_sheet
    
    for row_idx in range(10, max_row + 1):
        row_data = []
        row_profit = None
        
        # First find the profit for sorting later
        for header in headers:
            if header['text'] == 'Profit':
                profit_cell = ws.cell(row=row_idx, column=header['column'])
                if profit_cell.value and isinstance(profit_cell.value, (int, float)):
                    row_profit = float(profit_cell.value)
                break
        
        # Then extract all cell data
        for header in headers:
            col_idx = header['column']
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Special handling for block numbers
            is_block_number = header['text'] == 'block #'
            
            cell_data = {
                'text': get_cell_value(cell, is_block_number),
                'style': get_cell_style(cell),
                'hyperlink': get_cell_hyperlink(cell)
            }
            
            # Special handling for specific columns
            header_name = header['text']
            
            # For links like tx Hash, shorten display text but keep full link
            if header_name in ['tx Hash']:
                if cell_data['hyperlink'] and cell_data['text']:
                    if len(cell_data['text']) > 12:
                        # Save full hash but display shortened
                        cell_data['full_text'] = cell_data['text']
                        cell_data['text'] = cell_data['text'][:10] + '...'
            
            # Format Profit values with appropriate styling
            if header_name == 'Profit' and cell_data['text'] and cell_data['text'] != 'None':
                try:
                    value = float(cell_data['text'].replace(',', ''))
                    if value > 0:
                        cell_data['class'] = 'positive'
                    elif value < 0:
                        cell_data['class'] = 'negative'
                except ValueError:
                    pass
            
            # Add monospace class for code-like fields
            if header_name in ['block #', 'tx Hash', 'tx Index']:
                cell_data['class'] = cell_data.get('class', '') + ' monospace'
            
            row_data.append(cell_data)
        
        # Add profit for sorting
        data_rows.append({'profit': row_profit, 'data': row_data})
    
    # Sort rows by profit (descending)
    data_rows.sort(key=lambda x: x['profit'] if x['profit'] is not None else float('-inf'), reverse=True)
    
    # Extract only the row data after sorting
    sorted_rows = [row['data'] for row in data_rows]
    
    # Combine into main table data
    main_table_data = {
        'headers': headers,
        'rows': sorted_rows
    }
    
    return summary_data, main_table_data


def generate_html_report(excel_file: str, output_file: str = None, 
                         columns_to_keep: List[str] = None,
                         max_rows_per_sheet: int = 100,
                         entity_groups_dict: Dict[str, List[str]] = None) -> str:
    """
    Generate a single HTML file from an Excel blockchain analysis report.
    
    Args:
        excel_file: Path to the Excel file
        output_file: Path to save the HTML output (if None, return as string)
        columns_to_keep: List of column headers to include
        max_rows_per_sheet: Maximum number of rows to include per sheet (for size control)
        entity_groups_dict: Dictionary mapping primary entities to lists of sub-entities
    
    Returns:
        HTML content as string if output_file is None, otherwise None
    """
    if columns_to_keep is None:
        columns_to_keep = [
            'block #', 'tx Hash', 'tx Index', 'Balance', 'Profit', 
            'Bribe %', 'Profit Rank', 'Resim', 'Resim @ txIx 0', 
            'Resim @ txIx -1', 'miner'  # 'miner' at the end
        ]
    
    # Use the provided entity_groups dictionary if given
    global entity_groups
    if entity_groups_dict:
        entity_groups = entity_groups_dict
    
    wb = load_workbook(excel_file, data_only=True)  # data_only=True to get values instead of formulas
    
    # Group sheets by entity according to our dictionary
    reversed_entity_map = {}
    for primary_entity, sub_entities in entity_groups.items():
        # Create friendly names for tabs
        primary_friendly = get_friendly_name(primary_entity)
        
        for sub_entity in sub_entities:
            reversed_entity_map[sub_entity] = {
                'primary': primary_entity,
                'primary_friendly': primary_friendly
            }
    
    # Organize sheets by group
    grouped_sheets = defaultdict(list)
    
    # Extract data from each worksheet and organize by group
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Dashboard', 'Leaderboard']:
            continue  # Skip these tabs
            
        ws = wb[sheet_name]
        summary_data, main_table_data = extract_worksheet_data(ws, columns_to_keep, max_rows_per_sheet)
        
        if summary_data and main_table_data:  # Only add if data was successfully extracted
            sheets_data[sheet_name] = {
                'summary': summary_data,
                'main_table': main_table_data
            }
            
            # Get entity group based on first 8 characters
            primary_entity = get_entity_group(sheet_name)
            grouped_sheets[primary_entity].append(sheet_name)

            # Print debug information to see what's happening
            print("\nGroup assignments:")
            for group, sheets in grouped_sheets.items():
                print(f"Group: {group}")
                for sheet in sheets:
                    print(f"  - {sheet}")

            
            # Get entity group based on our mapping
            if sheet_name in reversed_entity_map:
                primary_entity = reversed_entity_map[sheet_name]['primary']
                grouped_sheets[primary_entity].append(sheet_name)
            else:
                # Use default grouping for any sheets not in our map
                group = get_entity_group(sheet_name)
                grouped_sheets[group].append(sheet_name)
    
    # Get timestamp for the report
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Generate HTML with grouped tabs
    html_content = generate_html_template(sheets_data, grouped_sheets, reversed_entity_map, timestamp)
    
    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        return None
    else:
        return html_content


def generate_html_template(sheets_data: Dict, grouped_sheets: Dict, 
                         entity_map: Dict, timestamp: str) -> str:
    """Generate the complete HTML template with nested tab navigation."""
    
    # Create a minimal HTML file with nested tab navigation
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="screen-orientation" content="any">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="HandheldFriendly" content="true">
    <title>Blockchain Analysis Report</title>
    <style>
        /* Base styles */
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, 'Open Sans', 'Helvetica Neue', sans-serif;
            margin: 0;
            padding: 0;
            color: #333;
            background-color: #f7f9fc;
            font-size: 14px;
        }}
        .container {{
            max-width: 100%;
            padding: 15px;
            margin: 0 auto;
        }}
        .report-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }}
        .report-title {{
            font-size: 20px;
            font-weight: bold;
            margin: 0;
            color: #2c3e50;
        }}
        .timestamp {{
            font-size: 12px;
            color: #7f8c8d;
        }}
        
        /* Primary Tab Navigation */
        .primary-tabs {{
            display: flex;
            flex-wrap: nowrap;
            overflow-x: auto;
            background-color: #fff;
            border-radius: 5px 5px 0 0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 0;
        }}
        .primary-tab {{
            padding: 12px 20px;
            cursor: pointer;
            transition: background-color 0.3s;
            font-weight: bold;
            white-space: nowrap;
            border-bottom: 3px solid transparent;
        }}
        .primary-tab.active {{
            background-color: #3498db;
            color: white;
            border-bottom-color: #2980b9;
        }}
        .primary-tab:hover:not(.active) {{
            background-color: #f1f1f1;
            border-bottom-color: #ddd;
        }}
        
        /* Secondary Tab Navigation */
        .secondary-tabs {{
            display: flex;
            flex-wrap: nowrap;
            overflow-x: auto;
            background-color: #f1f1f1;
            border-radius: 0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 15px;
        }}
        .secondary-tab {{
            padding: 8px 15px;
            cursor: pointer;
            transition: background-color 0.3s;
            font-weight: 500;
            white-space: nowrap;
            font-size: 13px;
        }}
        .secondary-tab.active {{
            background-color: #fff;
            color: #3498db;
            border-bottom: 2px solid #3498db;
        }}
        .secondary-tab:hover:not(.active) {{
            background-color: #e5e5e5;
        }}
        
        /* Content Containers */
        .primary-content, .secondary-content {{
            display: none;
        }}
        .primary-content.active, .secondary-content.active {{
            display: block;
        }}
        
        /* Summary section styles */
        .summary-section {{
            background-color: white;
            border-radius: 5px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            padding: 15px;
            margin-bottom: 15px;
        }}
        .wallet-address {{
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 15px;
            word-break: break-all;
            background-color: #f8f9fa;
            padding: 8px;
            border-radius: 4px;
            border-left: 4px solid #3498db;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
        }}
        .summary-card {{
            background-color: #f8f9fa;
            border-radius: 5px;
            padding: 12px;
        }}
        .summary-row {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 8px;
            border-bottom: 1px solid #eee;
            padding-bottom: 8px;
        }}
        .summary-row:last-child {{
            margin-bottom: 0;
            border-bottom: none;
            padding-bottom: 0;
        }}
        .summary-label {{
            font-weight: 500;
            color: #7f8c8d;
        }}
        .summary-value {{
            font-weight: 600;
            color: #2c3e50;
        }}
        
        /* Transaction table styles */
        .table-section {{
            background-color: white;
            border-radius: 5px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            padding: 15px;
            margin-bottom: 15px;
            overflow-x: auto;
        }}
        
        /* DataTables overrides for better alignment */
        table.dataTable {{
            width: 100% !important;
            margin: 0 !important;
        }}
        table.dataTable thead th, 
        table.dataTable tbody td {{
            padding: 8px 10px;
        }}
        table.dataTable thead th {{
            background-color: #f8f9fa;
            border-bottom: 2px solid #ddd;
        }}
        table.dataTable tbody td {{
            border-bottom: 1px solid #f2f2f2;
        }}
        .dataTables_wrapper .dataTables_filter input {{
            margin-left: 5px;
        }}
        
        /* Additional styles for specific elements */
        .positive {{
            color: #27ae60 !important;
        }}
        .negative {{
            color: #e74c3c !important;
        }}
        .monospace {{
            font-family: 'SFMono-Regular', Consolas, 'Liberation Mono', Menlo, monospace;
        }}
        
        /* Mobile optimizations */
        @media screen and (max-width: 768px) {{
            .container {{
                padding: 10px;
            }}
            .table-section {{
                padding: 10px;
            }}
            .primary-tab, .secondary-tab {{
                padding: 10px;
            }}
            .dataTables_wrapper .dataTables_filter input {{
                width: 80px;
            }}
        }}
        
        /* Landscape optimization */
        @media screen and (orientation: landscape) {{
            .summary-grid {{
                grid-template-columns: repeat(3, 1fr);
            }}
            .primary-tabs, .secondary-tabs {{
                flex-wrap: wrap;
            }}
        }}
    </style>
    <!-- DataTables minimal CSS -->
    <link rel="stylesheet" href="https://cdn.datatables.net/1.11.5/css/jquery.dataTables.min.css">
</head>
<body>
    <div class="container">
        <div class="report-header">
            <h1 class="report-title">Blockchain Transaction Analysis Report</h1>
            <div class="timestamp">Generated: {timestamp}</div>
        </div>
        
        <!-- Primary Tabs (Entity Groups) -->
        <div class="primary-tabs">
'''
    
    # Generate primary tabs for each entity group
    for i, (group_name, sheet_names) in enumerate(grouped_sheets.items()):
        active_class = 'active' if i == 0 else ''
        # Use friendly name if available, otherwise shorten the address
        display_name = get_friendly_name(group_name)
        html += f'            <div class="primary-tab {active_class}" data-group="{group_name}">{display_name}</div>\n'
    
    html += '''        </div>
        
'''
    
    # Generate content for each primary tab (entity group)
    for i, (group_name, sheet_names) in enumerate(grouped_sheets.items()):
        active_class = 'active' if i == 0 else ''
        
        html += f'''        <div id="{group_name}-content" class="primary-content {active_class}">
            <!-- Secondary Tabs (Individual Sheets) -->
            <div class="secondary-tabs">
'''
        
        # Generate secondary tabs for each sheet in this group
        for j, sheet_name in enumerate(sheet_names):
            active_class = 'active' if j == 0 else ''
            # Use friendly name if available, otherwise shorten the address
            display_name = get_friendly_name(sheet_name)
            html += f'                <div class="secondary-tab {active_class}" data-sheet="{sheet_name}">{display_name}</div>\n'
        
        html += '''            </div>
            
'''
        
        # Generate content for each secondary tab (individual sheet)
        for j, sheet_name in enumerate(sheet_names):
            active_class = 'active' if j == 0 else ''
            
            # Extract data for this sheet
            data = sheets_data[sheet_name]
            summary = data['summary']
            address = summary['address']
            summary_rows = summary['summary_rows']
            
            # Extract main table data
            main_table = data['main_table']
            headers = main_table['headers']
            rows = main_table['rows']
            
            html += f'''            <div id="{sheet_name}-content" class="secondary-content {active_class}">
                <div class="summary-section">
                    <div class="wallet-address">{address}</div>
                    <div class="summary-grid">
                        <div class="summary-card">
'''
            
            # Add first 2 summary rows
            for k, row in enumerate(summary_rows[:2]):
                label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
                value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
                
                # For start/end blocks, add monospace class
                if row['label'] in ['Start Block:', 'End Block:']:
                    value_style += ' font-family: monospace;'
                
                html += f'''                            <div class="summary-row">
                                <span class="summary-label" style="{label_style}">{row['label']}</span>
                                <span class="summary-value" style="{value_style}">{row['value']}</span>
                            </div>
'''
            
            html += '''                        </div>
                        <div class="summary-card">
'''
            
            # Add last 2 summary rows
            for k, row in enumerate(summary_rows[2:]):
                label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
                value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
                html += f'''                            <div class="summary-row">
                                <span class="summary-label" style="{label_style}">{row['label']}</span>
                                <span class="summary-value" style="{value_style}">{row['value']}</span>
                            </div>
'''
            
            html += '''                        </div>
                        <div class="summary-card">
'''
            
            # Add extra summary info (from columns G and H)
            for k, row in enumerate(summary_rows):
                label_style = ' '.join([f'{k}: {v};' for k, v in row['label_style'].items()])
                value_style = ' '.join([f'{k}: {v};' for k, v in row['value_style'].items()])
                html += f'''                            <div class="summary-row">
                                <span class="summary-label" style="{label_style}">{row['extra_label']}</span>
                                <span class="summary-value" style="{value_style}">{row['extra_value']}</span>
                            </div>
'''
            
            html += '''                        </div>
                    </div>
                </div>
                
                <div class="table-section">
                    <table id="''' + f"{sheet_name}-table" + '''" class="display nowrap compact stripe">
                        <thead>
                            <tr>
'''
            
            # Add table headers
            for header in headers:
                header_style = ' '.join([f'{k}: {v};' for k, v in header['style'].items()])
                # For block #, ensure monospace formatting
                header_classes = 'monospace' if header['text'] == 'block #' else ''
                html += f'''                                <th style="{header_style}" class="{header_classes}">{header['text']}</th>
'''
            
            html += '''                            </tr>
                        </thead>
                        <tbody>
'''
            
            # Add table rows
            for row in rows:
                html += '''                            <tr>
'''
                for cell in row:
                    cell_style = ' '.join([f'{k}: {v};' for k, v in cell['style'].items()])
                    cell_text = cell['text']
                    
                    # Combine classes
                    cell_classes = cell.get('class', '').strip()
                    
                    # Handle hyperlinks
                    if cell['hyperlink']:
                        # For tx hashes, use the full text in the hyperlink but display shortened
                        display_text = cell['full_text'] if 'full_text' in cell else cell_text
                        html += f'''                                <td style="{cell_style}" class="{cell_classes}"><a href="{cell['hyperlink']}" target="_blank">{display_text}</a></td>
'''
                    else:
                        html += f'''                                <td style="{cell_style}" class="{cell_classes}">{cell_text}</td>
'''
                
                html += '''                            </tr>
'''
            
            html += '''                        </tbody>
                    </table>
                </div>
            </div>
'''
        
        html += '''        </div>
'''
    
    # Add minimal JavaScript for interactivity
    html += '''        
    </div>

    <!-- Minimal required JavaScript -->
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.11.5/js/jquery.dataTables.min.js"></script>
    
    <script>
        $(document).ready(function() {
            // Primary tab switching
            $('.primary-tab').on('click', function() {
                const groupId = $(this).attr('data-group');
                
                // Update active primary tab
                $('.primary-tab').removeClass('active');
                $(this).addClass('active');
                
                // Show active primary content
                $('.primary-content').removeClass('active');
                $('#' + groupId + '-content').addClass('active');
                
                // Initialize tables in newly visible content if needed
                $('#' + groupId + '-content .secondary-content.active table.dataTable').each(function() {
                    $(this).DataTable().columns.adjust();
                });
            });
            
            // Secondary tab switching
            $('.secondary-tab').on('click', function() {
                const sheetId = $(this).attr('data-sheet');
                const parentGroup = $(this).closest('.primary-content');
                
                // Update active secondary tab within this primary tab
                parentGroup.find('.secondary-tab').removeClass('active');
                $(this).addClass('active');
                
                // Show active secondary content
                parentGroup.find('.secondary-content').removeClass('active');
                $('#' + sheetId + '-content').addClass('active');
                
                // Initialize table in newly visible content
                $('#' + sheetId + '-table').DataTable().columns.adjust();
            });
            
            // Initialize DataTables with minimal options
            $('table').each(function() {
                $(this).DataTable({
                    paging: true,
                    searching: true,
                    ordering: true,
                    pageLength: 25,
                    lengthMenu: [10, 25, 50, 100],
                    order: [[3, 'desc']], // Sort by Profit column (adjust index if needed)
                    scrollX: true,
                    deferRender: true,
                    language: {
                        search: "Filter:",
                        lengthMenu: "Show _MENU_",
                        info: "_START_ to _END_ of _TOTAL_"
                    },
                    dom: '<"top"lf>rt<"bottom"ip>',
                    columnDefs: [
                        // Set column-specific rendering options
                        {targets: 0, className: 'monospace'}, // Block number
                        {targets: 1, className: 'monospace'}, // TX hash
                    ]
                });
            });
            
            // Handle orientation changes
            window.addEventListener('orientationchange', function() {
                setTimeout(function() {
                    $('.primary-content.active .secondary-content.active table.dataTable').each(function() {
                        $(this).DataTable().columns.adjust();
                    });
                }, 200);
            });
            
            // Force tables to adjust after all content is loaded
            $(window).on('load', function() {
                setTimeout(function() {
                    $('.primary-content.active .secondary-content.active table.dataTable').each(function() {
                        $(this).DataTable().columns.adjust();
                    });
                }, 500);
            });
        });
    </script>
</body>
</html>'''

    return html


# Example usage
if __name__ == "__main__":
    # Generate HTML report from Excel file
    excel_file = "reports/MEV_daily_multiple_addresses_22016035_P&L.xlsx"
    html_output = "reports/blockchain_analysis_report.html"

    # Only keep these columns in the specified order
    columns_to_keep = [
        'block #', 'tx Hash', 'tx Index', 'miner',
        'Balance', 'Total Tip', 'Extractable Value', 'Profit',
        'Bribe %', 'Profit Rank', 'Resim', 'Resim @ txIx 0', 'Resim @ txIx -1'
    ]

    # Limit rows for size control (adjust as needed)
    max_rows_per_sheet = 100

    generate_html_report(excel_file, html_output, columns_to_keep, max_rows_per_sheet)
    print(f"HTML report generated: {html_output}")
