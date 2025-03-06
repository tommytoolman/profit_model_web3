import os
import time
import json
import math
import requests
import backoff
import logging
import asyncio
import aiohttp
import nest_asyncio
import subprocess
import tempfile
import pandas as pd
import numpy as np

from web3 import Web3
from decimal import Decimal
from eth_typing import HexStr
from hexbytes import HexBytes
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM
from datetime import datetime, timezone
from typing import Dict, List, Optional
from tqdm.notebook import tqdm
from pathlib import Path
from PIL import Image
from dotenv import load_dotenv

from lib_etherscan_funcs import get_address_tx_hashes_and_blocks
from chain_lib import w3_deejmon_http, deejmon_http, chain_data, minimal_abi
from metadata import miner_map, token_contracts, weth_contract, usdt_contract, usdc_contract, stETH_contract
from block_calculation_engine import get_timestamps, fetch_block_number, main

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
# from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment

load_dotenv()

W3 = w3_deejmon_http

etherscan_api_key = os.environ.get("etherscan_api_key")
address_list = os.environ.get("address_list").split(",")

selected_periods = ["current_datetime", "yesterday", "midnight", "yesterday_midnight", "start_of_week", "start_of_month"] # "start_of_1w"]

time_data = main(["Ethereum"], selected_periods[0:2])


def connect_with_retries(web3, retries=5, delay=0.1):
    for attempt in range(retries):
        if web3.is_connected():
            print("Connected to Ethereum node!")
            return True
        else:
            print(f"Connection attempt {attempt + 1} failed. Retrying in {delay} seconds...")
            time.sleep(delay)
    return False


def restructure_df(df):

    restructured_data = {
        'block_number': {},
        'datetime': {},
        'timestamp': {},
    }

    # For each time period (current_datetime, midnight, etc.)
    for period in df.columns:
        # Get the values for this period
        values = df[period].iloc[0]  # Since we only have one row (Ethereum)

        # Add values to our restructured data
        restructured_data['timestamp'][period] = values['timestamp']
        restructured_data['datetime'][period] = values['datetime']
        restructured_data['block_number'][period] = values['block_number']

    return pd.DataFrame(restructured_data).transpose()
